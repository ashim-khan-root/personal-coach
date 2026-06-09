"""Learn from real quotation XLSX files. Extract products, prices, and patterns.

Usage:
  py -3 coach/tools/learn_quotation.py <file.xlsx> [file2.xlsx ...]
  py -3 coach/tools/learn_quotation.py --dir <folder>
  py -3 coach/tools/learn_quotation.py --all-existing
"""
import sys, json, re, difflib
from pathlib import Path
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import openpyxl

COACH_DIR = Path(__file__).resolve().parent.parent
QUOTATION_DIRS = [
    COACH_DIR / "memory" / "quotations",
    COACH_DIR.parent / "quotation-coach" / "coach" / "memory" / "quotations",
]

# Which rates.json to update (quotation-coach has the more complete MOI version)
TARGET_RATES = COACH_DIR.parent / "quotation-coach" / "coach" / "memory" / "rates.json"
# Fallback to main coach
FALLBACK_RATES = COACH_DIR / "memory" / "rates.json"

# Patterns to categorize products
CATEGORY_PATTERNS = {
    "cameras": [
        r"\b\d+\s*MP", r"\bcamera\b", r"\bbullet\b", r"\beyeball\b",
        r"\bvarifocal\b", r"\bkpoi\b", r"\banpr\b", r"\blpr\b",
        r"\bthermal\b", r"\bdome\b",
    ],
    "nvrs": [
        r"\bnvr\b", r"\bnetwork.?video.?recorder\b", r"\brecorder\b",
    ],
    "hdd": [
        r"\bhdd\b", r"\bhard\s*disk\b", r"\bsurveillance\s*hdd\b",
        r"\b(?<!\d)(\d+)\s*TB\b", r"\bstorage\b",
    ],
    "switches": [
        r"\bswitch\b", r"\bpoe\b", r"\bgigabit\b",
    ],
    "racks": [
        r"\brack\b", r"\b(\d+)U\b.*rack", r"\bcabinet\b",
    ],
    "monitors": [
        r"\bmonitor\b", r"\bdisplay\b", r"\blcd\b", r"\bled\b",
        r'\b\d+(\.\d+)?"\s*(inch|led|monitor)',
    ],
    "workstation": [
        r"\bworkstation\b", r"\boptiplex\b", r"\bprecision\b",
        r"\bdell\b.*\bi[3579]\b", r"\bdesktop\b",
    ],
    "ups": [
        r"\bups\b", r"\buninterruptible\b", r"\bbattery\b",
        r"\b(?<!\d)(\d+)\s*KVA\b",
    ],
    "accessories": [
        r"\bmount\b", r"\bbracket\b", r"\bpatch\b", r"\bcable\b",
        r"\bconduit\b", r"\bpdu\b", r"\bpanel\b",
    ],
    "licenses": [
        r"\blicense\b", r"\bsoftware\b", r"\bvms\b",
    ],
    "services": [
        r"\binstallation\b", r"\bmaintenance\b", r"\bamc\b",
        r"\bdsa\b", r"\bdia\b", r"\bconfiguration\b", r"\bcommissioning\b",
    ],
}


def load_rates(path):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_rates(path, data):
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def detect_moi(filename):
    return bool(re.search(r"moi", filename, re.IGNORECASE))


def categorize_product(model, description):
    text = f"{model} {description}".lower()
    for cat, patterns in CATEGORY_PATTERNS.items():
        for pat in patterns:
            if re.search(pat, text, re.IGNORECASE):
                return cat
    return "other"


def generate_key(model, price):
    key = model.lower().strip()
    key = re.sub(r"[^a-z0-9\s]", "", key)
    key = re.sub(r"\s+", "_", key)[:40]
    if not key:
        key = f"item_{int(price)}"
    return key


def find_product_in_rates(rates, model, price):
    """Try to find a matching product across all categories."""
    for cat_key, cat_items in rates.items():
        if cat_key.startswith("_"):
            continue
        if isinstance(cat_items, dict):
            for pkey, pdata in cat_items.items():
                if isinstance(pdata, dict) and "price" in pdata:
                    if abs(pdata.get("price", 0) - price) < 5:
                        if pdata.get("model", "").lower() == model.lower():
                            return cat_key, pkey
                        # Check if model names are similar
                        m1 = pdata.get("model", "").lower()[:20]
                        m2 = model.lower()[:20]
                        if m1 and m2 and (m1 == m2 or m1 in m2 or m2 in m1):
                            return cat_key, pkey
    return None, None


def extract_items(xlsx_path):
    """Parse a quotation XLSX and extract line items."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    items = []
    customer = ""
    system_desc = ""
    date_str = ""
    header_row = None

    # Find header row (#, Model, Description, Qty, Unit Price, Amount)
    header_row = None
    header_row_num = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        vals = [str(v).strip().lower() if v else "" for v in row]
        if any("model" in v for v in vals) and any("description" in v for v in vals):
            header_row = row
            header_row_num = row_idx
            break

    if not header_row:
        return None, customer, system_desc, date_str

    # Extract metadata from header area
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        texts = [str(v).strip() if v else "" for v in row]
        full = " ".join(texts)
        m = re.search(r"Customer:\s*(.+)", full)
        if m:
            customer = m.group(1).strip()
        m = re.search(r"System:\s*(.+)", full)
        if m:
            system_desc = m.group(1).strip()
        m = re.search(r"Date:\s*([\d-]+)", full)
        if m:
            date_str = m.group(1)

    # Find column indices
    col_map = {}
    for i, v in enumerate(header_row):
        v = str(v).lower().strip() if v else ""
        if v in ("#", "no", "item"):
            col_map["num"] = i
        elif "model" in v or "product" in v or "description" in v:
            if "model" in v or "product" in v:
                col_map["model"] = i
            else:
                col_map["description"] = i
        elif "qty" in v or "quantity" in v:
            col_map["qty"] = i
        elif "unit" in v or "unit price" in v:
            col_map["unit_price"] = i
        elif "amount" in v or "total" in v:
            col_map["amount"] = i

    model_col = col_map.get("model", col_map.get("description", 2))
    desc_col = col_map.get("description", 2)
    qty_col = col_map.get("qty", 3)
    price_col = col_map.get("unit_price", 4)
    amount_col = col_map.get("amount", 5)

    # Parse line items (rows after header)
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        vals = list(row)
        if not vals or not vals[0]:
            continue
        first = str(vals[0]).strip()
        if first.lower() in ("subtotal", "grand total", "total", ""):
            break
        if not first.isdigit():
            continue

        model = str(vals[model_col]).strip() if model_col < len(vals) and vals[model_col] else ""
        desc = str(vals[desc_col]).strip() if desc_col < len(vals) and vals[desc_col] else ""
        qty = vals[qty_col] if qty_col < len(vals) and vals[qty_col] else 0
        unit_price = vals[price_col] if price_col < len(vals) and vals[price_col] else 0

        try:
            qty = int(float(str(qty)))
        except (ValueError, TypeError):
            qty = 1
        try:
            unit_price = float(str(unit_price).replace(",", ""))
        except (ValueError, TypeError):
            unit_price = 0

        if unit_price > 0 and model:
            items.append({
                "model": model,
                "description": desc[:200],
                "qty": qty,
                "unit_price": int(unit_price),
            })

    return items, customer, system_desc, date_str


def analyze_items(items, rates, is_moi):
    """Match extracted items to rates.json categories. Report new/mismatched products."""
    matched = []
    new_products = []
    category_totals = defaultdict(float)

    for item in items:
        cat, pkey = find_product_in_rates(rates, item["model"], item["unit_price"])
        item["matched_category"] = cat
        item["matched_key"] = pkey

        if cat and pkey:
            matched.append(item)
            category_totals[cat] += item["unit_price"] * item["qty"]
        else:
            cat = categorize_product(item["model"], item["description"])
            item["suggested_category"] = cat
            new_products.append(item)

    return matched, new_products, category_totals


def display_analysis(xlsx_name, items, customer, system_desc, is_moi,
                     matched, new_products, category_totals):
    print(f"\n{'='*60}")
    print(f"FILE: {xlsx_name}")
    print(f"{'MOI COMPLIANT' if is_moi else 'STANDARD'}")
    print(f"Customer: {customer}")
    print(f"System: {system_desc}")
    print(f"{'='*60}")

    print(f"\nLine Items ({len(items)}):")
    for item in items:
        tag = "✓" if item.get("matched_category") else "✗ NEW"
        cat = item.get("matched_category") or item.get("suggested_category", "?")
        print(f"  {tag} [{cat:15s}] {item['model'][:50]:50s}  "
              f"x{item['qty']}  @ {item['unit_price']} QAR")

    if category_totals:
        total = sum(category_totals.values())
        print(f"\nMatched Total: {int(total):,} QAR")
        for cat, amt in sorted(category_totals.items(), key=lambda x: -x[1]):
            print(f"  {cat:15s}: {int(amt):,} QAR")

    if new_products:
        print(f"\n⚠ NEW PRODUCTS (not in rates.json):")
        for p in new_products:
            print(f"  [{p.get('suggested_category','?')}] {p['model']} — {p['unit_price']} QAR")
            print(f"    → {p['description'][:100]}")


def find_all_quotations():
    found = []
    for d in QUOTATION_DIRS:
        if d.exists():
            for f in sorted(d.glob("quotation-*.xlsx")):
                if not f.name.startswith("~$"):
                    found.append(f)
    return found


def update_rates_with_new(rates, new_products, dry_run=True):
    """Suggest adding new products to rates.json."""
    updates = []
    for p in new_products:
        cat = p.get("suggested_category", "other")
        key = generate_key(p["model"], p["unit_price"])
        entry = {
            "model": p["model"],
            "description": p["description"][:200],
            "price": p["unit_price"],
        }
        if cat == "cameras":
            entry["moi_compliant"] = True
            entry["default_mount"] = "wall_mount"
            mp_match = re.search(r"(\d+)\s*MP", p["model"], re.IGNORECASE)
            entry["mp"] = int(mp_match.group(1)) if mp_match else 2
        updates.append((cat, key, entry))

    if not updates:
        return rates, updates

    if dry_run:
        print(f"\nProposed additions to rates.json:")
        for cat, key, entry in updates:
            print(f"  + rates['{cat}']['{key}'] = {entry['model']} ({entry['price']} QAR)")
    else:
        for cat, key, entry in updates:
            if cat not in rates:
                rates[cat] = {}
            if key not in rates[cat]:
                rates[cat][key] = entry
                print(f"  Added: rates['{cat}']['{key}']")
        print(f"  rates.json updated.")

    return rates, updates


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Learn from real quotation XLSX files")
    parser.add_argument("files", nargs="*", help="One or more XLSX quotation files")
    parser.add_argument("--dir", help="Directory containing XLSX quotation files")
    parser.add_argument("--all-existing", action="store_true",
                        help="Scan all existing quotation files in both projects")
    parser.add_argument("--dry-run", action="store_true", default=True,
                        help="Show what would be updated without writing (default)")
    parser.add_argument("--write", action="store_true",
                        help="Actually update rates.json")
    parser.add_argument("--target", choices=["quotation-coach", "coach"],
                        default="quotation-coach",
                        help="Which rates.json to update")
    args = parser.parse_args()

    # Collect files
    xlsx_files = []
    if args.all_existing:
        xlsx_files = find_all_quotations()
        print(f"Found {len(xlsx_files)} existing quotation files")
    if args.dir:
        d = Path(args.dir)
        if d.exists():
            xlsx_files.extend(sorted(d.glob("quotation-*.xlsx")))
    for f in args.files:
        p = Path(f)
        if p.exists():
            xlsx_files.append(p)

    if not xlsx_files:
        print("No quotation files found.")
        print("Usage: py -3 coach/tools/learn_quotation.py <file.xlsx> [--all-existing]")
        sys.exit(1)

    # Load target rates
    rates_path = TARGET_RATES if args.target == "quotation-coach" else FALLBACK_RATES
    if not rates_path.exists():
        rates_path = FALLBACK_RATES
    rates = load_rates(rates_path)
    print(f"Loaded rates from: {rates_path}")
    print(f"  Categories: {', '.join(k for k in rates if not k.startswith('_'))}")

    all_new_products = []

    for fp in xlsx_files:
        fname = fp.name
        if fname.startswith("~$"):
            continue
        is_moi = detect_moi(fname)
        result = extract_items(fp)
        if result is None:
            print(f"  Skipping {fname}: could not parse table structure")
            continue
        items, customer, system_desc, date_str = result
        if not items:
            print(f"  Skipping {fname}: no line items found")
            continue

        matched, new_products, cat_totals = analyze_items(items, rates, is_moi)
        display_analysis(fname, items, customer, system_desc, is_moi,
                         matched, new_products, cat_totals)
        all_new_products.extend(new_products)

    # Update rates if requested
    if all_new_products:
        rates, updates = update_rates_with_new(rates, all_new_products,
                                                dry_run=not args.write)
        if args.write and updates:
            save_rates(rates_path, rates)
            print(f"\n✓ rates.json saved to {rates_path}")
    else:
        print(f"\n✓ All products already known in rates.json")

    total_learned = sum(1 for p in all_new_products if p.get("unit_price", 0) > 0)
    print(f"\nSummary: {len(xlsx_files)} files processed, {total_learned} new products discovered")
    if total_learned > 0 and not args.write:
        print(f"Tip: re-run with --write to save new products to rates.json")


if __name__ == "__main__":
    main()
