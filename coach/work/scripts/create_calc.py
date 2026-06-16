from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Load Calculation"

# Column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 18

# Styles
title_font = Font(size=16, bold=True, color='FFFFFF')
header_font = Font(size=11, bold=True, color='FFFFFF')
data_font = Font(size=11)
title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
total_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws.merge_cells('A1:H1')
ws['A1'] = 'CCTV System Power Load Calculation'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

# Subtitle
ws.merge_cells('A2:H2')
ws['A2'] = 'Project: ___________________________________   Date: _______________'
ws['A2'].font = Font(size=11, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')
ws.row_dimensions[2].height = 25

# Headers
headers = ['#', 'Device Description', 'Qty', 'Unit Power (W)', 'Total Power (W)', 'Voltage (V)', 'Current (A)', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[3].height = 30

# Sample data
devices = [
    ('Bullet Camera 2MP', 8, 12, 0, '220V', '', 'Indoor/Outdoor'),
    ('Bullet Camera 5MP', 4, 15, 0, '220V (PoE)', '', 'Outdoor IR'),
    ('Dome Camera 2MP', 6, 10, 0, '220V (PoE)', '', 'Indoor'),
    ('PTZ Camera', 2, 35, 0, '220V', '', 'Outdoor'),
    ('NVR 16-Channel', 1, 60, 0, '220V', '', 'With HDD'),
    ('NVR 32-Channel', 1, 80, 0, '220V', '', 'With HDD'),
    ('PoE Switch 8-Port', 2, 65, 0, '220V', '', ''),
    ('PoE Switch 16-Port', 1, 120, 0, '220V', '', ''),
    ('UPS 1500VA', 1, 900, 0, '220V', '', 'Backup power'),
    ('CCTV Monitor 22"', 1, 45, 0, '220V', '', ''),
    ('Cabling & Accessories', 1, 50, 0, '220V', '', 'Estimated'),
]

for i, (device, qty, unit_power, total, voltage, current, notes) in enumerate(devices, 4):
    row = i
    total_power = qty * unit_power
    
    ws.cell(row=row, column=1, value=i-3).font = data_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2, value=device).font = data_font
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3, value=qty).font = data_font
    ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=3).border = thin_border
    
    ws.cell(row=row, column=4, value=unit_power).font = data_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=4).border = thin_border
    
    ws.cell(row=row, column=5, value=total_power).font = data_font
    ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=5).border = thin_border
    
    ws.cell(row=row, column=6, value=voltage).font = data_font
    ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=6).border = thin_border
    
    total_current = round(total_power / 220, 2) if total_power else 0
    ws.cell(row=row, column=7, value=total_current).font = data_font
    ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=7).number_format = '0.00'
    ws.cell(row=row, column=7).border = thin_border
    
    ws.cell(row=row, column=8, value=notes).font = data_font
    ws.cell(row=row, column=8).border = thin_border

# Total row
total_row = 4 + len(devices)
ws.cell(row=total_row, column=1, value='').border = thin_border
ws.cell(row=total_row, column=1).fill = total_fill
ws.cell(row=total_row, column=2, value='TOTAL').font = Font(size=12, bold=True)
ws.cell(row=total_row, column=2).fill = total_fill
ws.cell(row=total_row, column=2).border = thin_border
ws.cell(row=total_row, column=3, value=sum(d[1] for d in devices)).font = Font(size=12, bold=True)
ws.cell(row=total_row, column=3).fill = total_fill
ws.cell(row=total_row, column=3).alignment = Alignment(horizontal='center')
ws.cell(row=total_row, column=3).border = thin_border
ws.cell(row=total_row, column=4).fill = total_fill
ws.cell(row=total_row, column=4).border = thin_border
ws.cell(row=total_row, column=5, value=sum(d[1]*d[2] for d in devices)).font = Font(size=12, bold=True)
ws.cell(row=total_row, column=5).fill = total_fill
ws.cell(row=total_row, column=5).alignment = Alignment(horizontal='center')
ws.cell(row=total_row, column=5).border = thin_border
ws.cell(row=total_row, column=6).fill = total_fill
ws.cell(row=total_row, column=6).border = thin_border
ws.cell(row=total_row, column=7, value=round(sum(d[1]*d[2] for d in devices) / 220, 2)).font = Font(size=12, bold=True)
ws.cell(row=total_row, column=7).fill = total_fill
ws.cell(row=total_row, column=7).alignment = Alignment(horizontal='center')
ws.cell(row=total_row, column=7).border = thin_border
ws.cell(row=total_row, column=8, value='Total Load').font = Font(size=12, bold=True)
ws.cell(row=total_row, column=8).fill = total_fill
ws.cell(row=total_row, column=8).border = thin_border

# Notes section
note_row = total_row + 2
ws.merge_cells(f'A{note_row}:H{note_row}')
ws.cell(row=note_row, column=1, value='Notes:').font = Font(size=11, bold=True)

notes_text = [
    '1. All power calculations are approximate and should be verified by a qualified electrician.',
    '2. UPS sizing should consider total load + 30% headroom.',
    '3. Cable sizing must comply with local electrical codes (Qatar: QCS 2014 / Kahramaa).',
    '4. Consider future expansion when sizing power supply and UPS.',
    '5. For outdoor installations, ensure appropriate IP rating and surge protection.',
]
for j, note in enumerate(notes_text):
    r = note_row + 1 + j
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(row=r, column=1, value=note).font = Font(size=10, italic=True)

# Print settings
ws.sheet_properties.pageSetUpPr = None
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A4

# Save
output_path = 'calc-load_decoded.xlsx'
wb.save(output_path)
print(f'Created: {output_path}')
