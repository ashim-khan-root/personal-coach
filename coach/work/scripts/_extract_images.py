"""Extract all images from demo-template.xlsx and map them to rows"""
import openpyxl
from pathlib import Path
from openpyxl.drawing.image import Image

OUT_DIR = Path("coach/memory/images")
OUT_DIR.mkdir(parents=True, exist_ok=True)
QUOTE_OUT_DIR = Path("quotation-coach/coach/memory/images")
QUOTE_OUT_DIR.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook("demo-template.xlsx")
ws = wb["Sheet1"]

print(f"Total images found: {len(ws._images)}")
print()

# Build a map of row -> images based on anchor
image_map = {}  # row -> list of (col, image_data, ext)
for img in ws._images:
    anchor = img.anchor
    row = None
    col = None
    if hasattr(anchor, "_from"):
        row = anchor._from.row + 1  # 0-indexed -> 1-indexed
        col = anchor._from.col + 1
    elif hasattr(anchor, "row"):
        row = anchor.row
        col = anchor.col
    
    raw = img._data()
    if callable(raw):
        img_data = raw()
    else:
        img_data = raw
    fmt = img.format if hasattr(img, 'format') and img.format else 'png'
    
    if row:
        if row not in image_map:
            image_map[row] = []
        image_map[row].append((col, img_data, fmt))

# Print image locations with their corresponding item data
for r in sorted(image_map.keys()):
    items = ws.cell(row=r, column=2).value or ""
    model = ws.cell(row=r, column=1).value or ""
    qty = ws.cell(row=r, column=4).value or ""
    price = ws.cell(row=r, column=5).value or ""
    
    # Save each image
    for idx, (col, img_data, fmt) in enumerate(image_map[r]):
        ext = fmt.lower() if fmt else "png"
        safe_name = f"row{r:02d}-col{col}"
        filename = f"{safe_name}.{ext}"
        filepath = OUT_DIR / filename
        with open(filepath, "wb") as f:
            if hasattr(img_data, 'read'):
                f.write(img_data.read())
            else:
                f.write(img_data)
        
        # Also copy to quotation-coach
        qfilepath = QUOTE_OUT_DIR / filename
        with open(qfilepath, "wb") as f:
            if hasattr(img_data, 'read'):
                f.write(img_data.read())
            else:
                f.write(img_data)
        
        info = f"  Row {r} Col {col}: {filename}"
        if model or items:
            info += f" | {model or ''} {str(items)[:60]}"
        if qty or price:
            info += f" | Qty={qty} Price={price}"
        print(info)

print(f"\nTotal images extracted: {sum(len(v) for v in image_map.values())}")
print(f"Saved to: {OUT_DIR}")
