"""
STARFOX AI Quotation Maker v2 — MOI-Compliant CCTV & Security Quotes

Usage:
  python make_quotation.py "8 cameras 2MP"
  python make_quotation.py "3 cameras 4MP KPOI for mosque" --moi
  python make_quotation.py "2 cameras 4MP ANPR parking" --customer "Gate 1"
  python make_quotation.py "20 cameras 2MP for school" --moi --discount 10
  python make_quotation.py "thermal" --customer "Industrial Zone"
  python make_quotation.py --interactive
  python make_quotation.py --list-rates
  python make_quotation.py --update-rates <file.xlsx>
"""
import sys, json, re, datetime as dt
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
MEM_DIR = BASE / "memory"
RATES_PATH = MEM_DIR / "rates.json"
MOI_SPECS_PATH = MEM_DIR / "moi_specs.json"
QUOTES_DIR = MEM_DIR / "quotations"
QUOTES_DIR.mkdir(parents=True, exist_ok=True)
IMAGES_DIR = BASE / "memory" / "images"

CURRENCY = "QAR"
COMPANY_NAME = "STARFOX SECURITY"
COMPANY_INFO = "Salwa Road, Doha, Qatar | Tel: +974 44691741 | info@starfoxsecu.com | www.starfoxsecu.com"

TERMS_STANDARD = [
    "All the above prices are in QAR (Qatari Riyals)",
    "Terms of payment: 50% upon signing the contract as downpayment and 50% before installation and handover",
    "Warranty: 2 YEARS for NVR and Cameras",
    "Any changes or addition of cameras as recommended will be quoted separately",
    "Any civil, mechanical, and electrical work if any will be charged accordingly",
    "Materials prices may vary at the time of installation due to current regional situation",
    "This quotation is subject to STARFOX standard terms and conditions which are available on request",
    "Price and specifications may change without prior notice",
]

TERMS_MOI = [
    "All the above prices are in QAR (Qatari Riyals)",
    "Terms of payment: 50% upon signing the contract, 50% before installation and handover",
    "Warranty: 2 YEARS for NVR and Cameras",
    "System designed and installed per MOI SSD requirements (Law No. 9 of 2011)",
    "Storage retention: 120 days continuous recording with RAID 5 configuration",
    "System includes minimum 1-hour UPS backup as per MOI requirements",
    "Annual Maintenance Contract (AMC) is mandatory for MOI compliance",
    "Any changes or additions must be approved via MOI SSD portal",
    "Any civil, mechanical, and electrical work if any will be charged accordingly",
    "Materials prices may vary at the time of installation due to current regional situation",
    "This quotation is subject to STARFOX standard terms and conditions",
    "Price and specifications may change without prior notice",
]

# ── Arabic Translations ──────────────────────────────────────────────────
ARABIC = {
    "company_name": "ستار فوكس للأمن",
    "subtitle": "أنظمة كاميرات المراقبة والأمن المنزلي الذكي",
    "moi_subtitle": "— متوافقة مع القانون رقم 9 لسنة 2011 لوزارة الداخلية",
    "quotation": "عرض سعر",
    "date": "التاريخ",
    "customer": "العميل",
    "system": "النظام",
    "retention_days": "أيام التخزين",
    "raid": "RAID 5",
    "yes": "نعم",
    "no": "لا",
    "hdr_model": "الموديل",
    "hdr_desc": "الوصف",
    "hdr_qty": "الكمية",
    "hdr_unit_price": "سعر الوحدة (ريال)",
    "hdr_amount": "الإجمالي (ريال)",
    "subtotal": "المجموع",
    "discount": "الخصم",
    "grand_total": "الإجمالي النهائي",
    "terms_title": "الشروط والأحكام:",
    "company_info": "سلوى، الدوحة، قطر | هاتف: 974 44691741+ | info@starfoxsecu.com",
    "terms_moi": [
        "جميع الأسعار أعلاه بالريال القطري (QAR)",
        "شروط الدفع: 50% عند توقيع العقد و 50% قبل التركيب والتسليم",
        "الضمان: سنتان لجهاز التسجيل والكاميرات",
        "تم تصميم النظام وتركيبه حسب متطلبات وزارة الداخلية (القانون رقم 9 لسنة 2011)",
        "فترة التخزين: 120 يوماً تسجيل مستمر مع نظام RAID 5",
        "النظام مزود ببطارية احتياطية (UPS) لمدة ساعة كحد أدنى حسب متطلبات وزارة الداخلية",
        "عقد الصيانة السنوية (AMC) إلزامي للامتثال لوزارة الداخلية",
        "أي تغييرات أو إضافات يجب اعتمادها عبر بوابة وزارة الداخلية",
        "أي أعمال مدنية أو ميكانيكية أو كهربائية إن وجدت تحسب بشكل منفصل",
        "أسعار المواد قد تتغير وقت التركيب حسب الظروف الإقليمية",
        "عرض السعر هذا يخضع للشروط والأحكام القياسية لشركة ستار فوكس",
        "قد تتغير الأسعار والمواصفات دون إشعار مسبق",
    ],
    "terms_standard": [
        "جميع الأسعار أعلاه بالريال القطري (QAR)",
        "شروط الدفع: 50% عند توقيع العقد كدفعة مقدمة و 50% قبل التركيب والتسليم",
        "الضمان: سنتان لجهاز التسجيل والكاميرات",
        "أي تغييرات أو إضافة كاميرات كما يوصى بها تحسب بشكل منفصل",
        "أي أعمال مدنية أو ميكانيكية أو كهربائية إن وجدت تحسب بشكل منفصل",
        "أسعار المواد قد تتغير وقت التركيب حسب الظروف الإقليمية",
        "عرض السعر هذا يخضع للشروط والأحكام القياسية لشركة ستار فوكس",
        "قد تتغير الأسعار والمواصفات دون إشعار مسبق",
    ],
}
# Map English model names (substring) → Arabic model names
MODEL_AR = {
    "2MP IR Fixed Bullet": "كاميرا رصاصية ثابتة 2 ميجابكسل مع أشعة تحت حمراء",
    "2MP Lite IR Eyeball": "كاميرا عين سمكة 2 ميجابكسل مع أشعة تحت حمراء",
    "2MP AcuSense Varifocal": "كاميرا متغيرة البعد البؤري 2 ميجابكسل أكيو سينس",
    "4MP KPOI Camera": "كاميرا KPOI 4 ميجابكسل مع عدسة",
    "4MP ANPR/LPR Bullet": "كاميرا ANPR/LPR للوحات المركبات 4 ميجابكسل",
    "4MP IR Bullet": "كاميرا رصاصية 4 ميجابكسل مع أشعة تحت حمراء",
    "8MP 4K IR Bullet": "كاميرا رصاصية 8 ميجابكسل 4K مع أشعة تحت حمراء",
    "Secuview 2MP Outdoor Fixed Bullet IP Camera": "كاميرا سيكيوفي رصاصية خارجية 2 ميجابكسل IP بعدسة ثابتة",
    "Secuview 8MP Full HD Outdoor IP Camera": "كاميرا سيكيوفي خارجية 8 ميجابكسل كاملة الدقة IP مقاومة للماء",
    "Secuview 6MP Outdoor IP PTZ Camera": "كاميرا سيكيوفي PTZ خارجية 6 ميجابكسل IP مع عدسة أوتوفوكاس",
    "Secuview 12MP Outdoor IP Camera": "كاميرا سيكيوفي خارجية 12 ميجابكسل IP مع كشف ذكي",
    "Secuview WiFi Outdoor Weatherproof Security Camera": "كاميرا سيكيوفي خارجية لاسلكية WiFi مقاومة للعوامل الجوية",
    "Secuview Wireless IP Solar-powered": "كاميرا سيكيوفي لاسلكية IP تعمل بالطاقة الشمسية 4G",
    "Secuview 4G Smart Dash Cam": "كاميرا سيارة ذكية 4G سيكيوفي معتمدة من وزارة الداخلية",
    "Thermal Bi-Spectrum": "كاميرا حرارية ثنائية الطيف",
    "Channel 4K NVR": "جهاز تسجيل NVR لـ",
    "Channel NVR": "جهاز تسجيل NVR لـ",
    "16CH NVR for KPOI": "جهاز تسجيل NVR لـ KPOI 16 قناة",
    "16CH NVR for ANPR": "جهاز تسجيل NVR لـ ANPR 16 قناة",
    "Secuview 16CH 4K Hybrid NVR PoE Built-in": "جهاز تسجيل سيكيوفي NVR هجين 16 قناة 4K مزود بـ PoE",
    "Secuview 8CH 4K NVR Ultra HD PoE Built-in": "جهاز تسجيل سيكيوفي NVR 8 قناة 4K فائق الدقة مزود بـ PoE",
    "Secuview 8CH 4K NVR Non-PoE": "جهاز تسجيل سيكيوفي NVR 8 قناة 4K بدون PoE",
    "Secuview 16CH 4K Non-PoE Hybrid NVR": "جهاز تسجيل سيكيوفي NVR هجين 16 قناة 4K بدون PoE",
    "Surveillance HDD": "قرص صلب للمراقبة",
    "TB Surveillance HDD": "تيرابايت قرص صلب للمراقبة",
    "PoE+ Switch": "سويتش PoE+",
    "Port Gigabit": "منفذ جيجابت",
    "Secuview 4-Port Gigabit PoE Switch": "سويتش سيكيوفي 4 منافذ جيجابت PoE",
    "Secuview 8-Port Gigabit PoE+ Switch": "سويتش سيكيوفي 8 منافذ جيجابت PoE+",
    "Secuview 16-Port Gigabit PoE Switch": "سويتش سيكيوفي 16 منفذ جيجابت PoE",
    "Secuview 24-Port Gigabit PoE Switch": "سويتش سيكيوفي 24 منفذ جيجابت PoE",
    "Secuview 5-Port Gigabit Ethernet Switch": "سويتش إيثرنت سيكيوفي 5 منافذ جيجابت",
    "Secuview 8-Port Gigabit Ethernet Switch": "سويتش إيثرنت سيكيوفي 8 منافذ جيجابت",
    "Secuview 5-Port Fast Ethernet Switch": "سويتش إيثرنت سيكيوفي 5 منافذ سريع",
    "Secuview 8-Port Fast Ethernet Switch": "سويتش إيثرنت سيكيوفي 8 منافذ سريع",
    "Secuview 11ac 1200Mbps 4G LTE Wi-Fi Router": "راوتر سيكيوفي 4G LTE 1200 ميجابايت في الثانية مع واي فاي",
    "Secuview Outdoor Wireless 4G Router": "راوتر سيكيوفي خارجي لاسلكي 4G مقاوم للعوامل الجوية",
    "Wall-Mount Rack": "خزانة حائط",
    "Floor Mount Rack": "خزانة أرضية",
    "Secuview 19\" 4U SPCC Network Cabinet": "خزانة شبكات سيكيوفي 4U SPCC",
    "Secuview 19\" 9U SPCC Data Center Cabinet": "خزانة مركز بيانات سيكيوفي 9U SPCC",
    "Secuview 19\" 12U SPCC Network Cabinet": "خزانة شبكات سيكيوفي 12U SPCC",
    "Secuview 19\" 15U SPCC Network Cabinet": "خزانة شبكات سيكيوفي 15U SPCC",
    "Secuview 19\" 18U SPCC Network Cabinet": "خزانة شبكات سيكيوفي 18U SPCC",
    "Secuview 19\" 22U 600x600 SPCC Network Rack": "رف شبكات سيكيوفي 22U 600x600 SPCC",
    "Secuview 19\" 22U SPCC Network Rack 600x800": "رف شبكات سيكيوفي 22U 600x800 SPCC",
    "Full HD LED Monitor": "شاشة LED كاملة الدقة",
    "Secuview 22-Inch HD LED Monitor": "شاشة سيكيوفي LED 22 بوصة عالية الدقة",
    "inch 4K UHD Monitor": "بوصة 4K UHD شاشة",
    "Dell OptiPlex CCTV Workstation": "محطة عمل ديل أوبتي بلكس للمراقبة",
    "Dell Precision CCTV Workstation": "محطة عمل ديل بريسيجن للمراقبة",
    "Rack-Type Online UPS": "UPS اونلاين من نوع الرف",
    "Wall Mount Camera Bracket": "حامل تثبيت حائط للكاميرا",
    "Bullet Camera Base Box": "صندوق قاعدة للكاميرا الرصاصية",
    "Hanging/Pendant Mount Stand": "حامل تعليق للكاميرات",
    "Pole Mount Bracket": "حامل تثبيت على عمود",
    "Meter Galvanized Pole": "متر عمود مجلفن",
    "CAT6 UTP Patch Cord": "كابل باتش CAT6",
    "Rack PDU": "وحدة توزيع طاقة للرف",
    "CAT6 Patch Panel": "بانل باتش CAT6",
    "Horizontal Cable Manager": "مدير كابلات أفقي",
    "CAT6 UTP Cable": "كابل CAT6 UTP",
    "PVC Conduit": "ماسورة PVC",
    "HIK-Central VMS Base": "ترخيص HIK-Central VMS أساسي",
    "HIK-Central P-ANPR": "ترخيص HIK-Central P-ANPR",
    "HIK-Central Extra": "ترخيص إضافي HIK-Central",
    "VMS Software License": "رخصة برنامج إدارة الفيديو",
    "Camera Installation & Cabling": "تركيب الكاميرات وتمديد الكابلات",
    "NVR Configuration & Setup": "إعداد وتكوين جهاز التسجيل NVR",
    "DSA/DIA": "رسومات DSA/DIA والتوثيق والموافقات",
    "Annual Maintenance Contract": "عقد الصيانة السنوية",
    "Access Control System": "نظام تحكم دخول",
    "UPS": "UPS",
    "Battery Cabinet": "خزانة بطاريات",
    "External Battery": "بطارية خارجية",
    "Patch Cord": "كابل باتش",
    # Door phones / doorbells
    "Secuview 10\" WiFi Smart Video DoorPhone": "جهاز اتصال باب فيديو سيكيوفي 10 بوصة WiFi ذكي",
    "Secuview Full HD Wi-Fi Smart Doorphone": "كاميرا باب سيكيوفي ذكية خارجية WiFi عالية الدقة",
    "Secuview IP Smart Video Doorbell": "جرس باب فيديو سيكيوفي IP ذكي",
    "Secuview Smart Wi-Fi Waterproof Video Doorbell": "جرس باب فيديو سيكيوفي WiFi ذكي مقاوم للماء",
    # Smart locks
    "Secuview Mobile Control Smart Door Lock": "قفل باب ذكي سيكيوفي مع تحكم عبر الجوال",
    "Secuview WiFi Smart Door Lock for Home Safety": "قفل باب WiFi ذكي سيكيوفي لسلامة المنزل",
    "Secuview Fingerprint Smart Lock with Remote Control": "قفل ذكي سيكيوفي ببصمة الأصبع مع تحكم عن بعد",
    "Secuview Smart Home Door Lock with Remote Control": "قفل باب منزل ذكي سيكيوفي مع تحكم عن بعد",
    "Secuview Smart Glass Door Lock Device for Office": "قفل باب زجاجي ذكي سيكيوفي للمكاتب",
    "Secuview Mobile Control Fingerprint Smart Lock": "قفل ذكي سيكيوفي ببصمة الأصبع مع تحكم عبر الجوال",
    "Secuview Keyless Entry Savvy Digital Lock": "قفل رقمي سيكيوفي ذكي بدون مفتاح",
    "Secuview Keyless Entry Smart Door Lock with Fingerprint": "قفل باب ذكي سيكيوفي بدون مفتاح مع بصمة",
    "Secuview Smart Technology Door Lock": "قفل باب سيكيوفي بتقنية ذكية",
    "Secuview Waterproof Smart Courtyard Gate Lock": "قفل بوابة سيكيوفي ذكي مقاوم للماء للفناء الخارجي",
    # Smart switches
    "Secuview 1-Gang WiFi Remote Control Smart Switch": "مفتاح ذكي سيكيوفي 1-مجموعة WiFi مع تحكم عن بعد",
    "Secuview 1-Gang Brushed Panel Smart Switch": "مفتاح ذكي سيكيوفي 1-مجموعة بلوح فرشاة",
    "Secuview 4-Gang WiFi Remote Control Smart Switch": "مفتاح ذكي سيكيوفي 4-مجموعات WiFi مع تحكم عن بعد",
    "Secuview 4-Gang Brushed Panel Smart Switch": "مفتاح ذكي سيكيوفي 4-مجموعات بلوح فرشاة",
    "Secuview Smart IR Remote Controller": "جهاز تحكم عن بعد سيكيوفي ذكي بالأشعة تحت الحمراء",
    # PABX & Telephones
    "Secuview 16-Channel PABX Advanced Communication": "مقسم هاتفي سيكيوفي 16 خط اتصال متطور",
    "Secuview Hybrid PABX 16 Channels": "مقسم هاتفي هجين سيكيوفي 16 خط",
    "Secuview Caller ID Telephone": "هاتف سيكيوفي مع معرف المتصل",
    # Cables
    "Secuview 100m RG59+2C Coaxial Cable": "كابل سيكيوفي محوري 100 متر RG59 مع طاقة",
    "Secuview 305m RG59+2C Coaxial Cable": "كابل سيكيوفي محوري 305 متر RG59 مع طاقة",
    "Secuview RG6/U Coaxial Cable 305M": "كابل سيكيوفي محوري RG6 305 متر",
    "Secuview 2.0 HDTV 1.5M Cable": "كابل سيكيوفي HDMI 2.0 بطول 1.5 متر",
    "Secuview 2.0 HDTV 5M Cable": "كابل سيكيوفي HDMI 2.0 بطول 5 متر",
    "Secuview 2.0 HDTV 10M Cable": "كابل سيكيوفي HDMI 2.0 بطول 10 متر",
    "Secuview 2.0 HDTV 20M Cable": "كابل سيكيوفي HDMI 2.0 بطول 20 متر",
    # Accessories (mounts, brackets, etc.)
    "Secuview Universal Wall Mount Metal Bracket": "حامل تثبيت حائط معدني سيكيوفي عالمي",
    "Secuview 1258ZJ L-Type Wall Mount Bracket Plastic": "حامل تثبيت حائط سيكيوفي بلاستيك نوع L",
    "Secuview 1258 ABS L-Type Wall Mount Bracket": "حامل تثبيت حائط سيكيوفي معدني نوع L",
    "Secuview 1-2M TS Adjustable Metal Bracket": "حامل معدني سيكيوفي قابل للتعديل 1-2 متر",
    "Secuview 14-Inch Universal PTZ Mount": "حامل PTZ سيكيوفي عالمي 14 بوصة",
    "Secuview Adjustable PTZ Metal Mount Bracket": "حامل PTZ معدني سيكيوفي قابل للتعديل",
    "Secuview 105mm Waterproof Metal Junction Box": "صندوق توصيل سيكيوفي معدني مقاوم للماء 105 مم",
    "Secuview 120mm Waterproof Metal Junction Box": "صندوق توصيل سيكيوفي معدني مقاوم للماء 120 مم",
    "Secuview Compact Indoor Metal Junction Box": "صندوق توصيل سيكيوفي معدني داخلي مضغوط",
    "Secuview 160mm Waterproof Metal Junction Box": "صندوق توصيل سيكيوفي معدني مقاوم للماء 160 مم",
    "Secuview 12V 2A AC/DC Plug Power Adapter": "محول طاقة سيكيوفي 12V 2A AC/DC",
    "Secuview 12V 2A Waterproof Power Adapter": "محول طاقة سيكيوفي 12V 2A مقاوم للماء",
    "Secuview High Performance 12V 10A Adapter": "محول طاقة سيكيوفي عالي الأداء 12V 10A",
    "Secuview 12V 20A Power Supply": "مزود طاقة سيكيوفي 12V 20A",
    "Secuview 12V 30A Heavy-Duty Power Supply": "مزود طاقة سيكيوفي ثقيل 12V 30A",
    "Secuview 1x8 HDMI Splitter 4K": "موزع HDMI سيكيوفي 1x8 4K",
    "Secuview 4K HDMI Splitter 4 Screens": "موزع HDMI سيكيوفي 4K لأربع شاشات",
    "Secuview 4K 60Hz HDMI Extender 60M CAT6": "موسع HDMI سيكيوفي 4K 60Hz لمسافة 60 متر CAT6",
    "Secuview 4K HDMI+USB KVM Extender 120M": "موسع KVM سيكيوفي HDMI+USB 4K لمسافة 120 متر",
    "Secuview 400m HD Signal Video Balun": "بالون فيديو سيكيوفي عالي الدقة 400 متر",
    "Secuview Gigabit PoE Extender Weatherproof": "موسع PoE جيجابت سيكيوفي مقاوم للعوامل الجوية",
    "Secuview USB RJ45 Extender 150ft": "موسع USB سيكيوفي عبر RJ45 لمسافة 150 قدم",
    "Secuview Mounting Metal Bracket for CCTV": "حامل تثبيت سيكيوفي معدني لكاميرات المراقبة",
}

def get_model_ar(model_en):
    for key, val in MODEL_AR.items():
        if key in model_en:
            return val
    return model_en

def load_rates():
    with open(RATES_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def load_moi_specs():
    with open(MOI_SPECS_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def save_rates(rates):
    with open(RATES_PATH, "w", encoding="utf-8") as f:
        json.dump(rates, f, ensure_ascii=False, indent=2)
    print(f"[rates] Saved to {RATES_PATH}")

def list_rates():
    rates = load_rates()
    print("="*60)
    print(f"  STARFOX RATE CARD ({CURRENCY})")
    print("="*60)
    for category, items in rates.items():
        if category.startswith("_"):
            continue
        print(f"\n  [{category.upper()}]")
        for key, item in items.items():
            p = item.get("price", 0)
            m = item.get("model", key)
            moi = " [MOI]" if item.get("moi_compliant") else ""
            print(f"    {key:25s} {p:>8,} {CURRENCY}  | {m}{moi}")
    print()

# ── Extra Items (non-CCTV products) ───────────────────────────────────────
# Order matters: more specific phrases must come before general ones.
EXTRA_ITEM_PATTERNS = [
    # Door phones / doorbells
    ("door phone monitor",     "doorphone",      "secuview_10in_wifi_kit"),
    ("door phone camera",      "doorphone",      "secuview_hd_wifi_camera"),
    ("video door phone",       "doorphone",      "secuview_10in_wifi_kit"),
    ("door phone",             "doorphone",      "secuview_10in_wifi_kit"),
    ("doorbell",               "doorphone",      "secuview_ip_doorbell"),
    ("video doorbell",         "doorphone",      "secuview_ip_doorbell"),
    # Smart locks
    ("smart lock",             "smart_locks",    "secuview_fingerprint_lock"),
    ("fingerprint lock",       "smart_locks",    "secuview_fingerprint_lock"),
    ("digital lock",           "smart_locks",    "secuview_keyless_entry"),
    ("gate lock",              "smart_locks",    "secuview_gate_lock"),
    ("glass door lock",        "smart_locks",    "secuview_glass_door_lock"),
    # Smart switches
    ("smart switch",           "smart_switches", "secuview_4gang_wifi"),
    ("smart socket",           "smart_switches", "secuview_1gang_wifi"),
    ("smart remote",           "smart_switches", "secuview_smart_ir_remote"),
    ("ir remote",              "smart_switches", "secuview_smart_ir_remote"),
    ("ir controller",          "smart_switches", "secuview_smart_ir_remote"),
    # Access points
    ("access point",           "access_points",  "secuview_wifi6_ax1800"),
    # TV / Satellite
    ("tv satellite outlet",    "tv_satellite",   "secuview_internet_sat_socket"),
    ("tv satellite",           "tv_satellite",   "secuview_internet_sat_socket"),
    ("internet satellite",     "tv_satellite",   "secuview_internet_sat_socket"),
    ("satellite outlet",       "tv_satellite",   "secuview_internet_sat_socket"),
    # Speakers
    ("wall speaker",           "speakers",       "secuview_20w_hanging"),
    ("ceiling speaker",        "speakers",       "secuview_10w_ceiling"),
    ("volume controller",      "volume_controls","secuview_30w_vc_steel"),
    ("volume control",         "volume_controls","secuview_30w_vc_steel"),
    ("speaker",                "speakers",       "secuview_10w_ceiling"),
    # PABX / Telephones
    ("pabx",                   "pabx",           "secuview_16ch_pabx"),
    ("telephone",              "pabx",           "secuview_caller_id_phone"),
    ("phone",                  "pabx",           "secuview_caller_id_phone"),
    # Cables
    ("hdmi cable",             "cables",         "secuview_hdmi_5m"),
    ("hdmi",                   "cables",         "secuview_hdmi_5m"),
    ("coaxial cable",          "cables",         "secuview_rg59_100m"),
    ("coax cable",             "cables",         "secuview_rg59_100m"),
    # Routers
    ("4g router",              "routers",        "secuview_4g_lte_1200"),
    ("lte router",             "routers",        "secuview_4g_lte_1200"),
    # Ethernet switches
    ("network switch",         "ethernet_switches","secuview_8port_gig"),
    ("gigabit switch",         "ethernet_switches","secuview_8port_gig"),
]

EXTRA_CAT_LABEL = {
    "doorphone":       "Door Phone",
    "smart_locks":     "Smart Locks",
    "smart_switches":  "Smart Switches",
    "access_points":   "Access Points",
    "tv_satellite":    "TV / Satellite",
    "speakers":        "Speakers",
    "volume_controls": "Volume Controls",
    "pabx":            "PABX & Telephones",
    "cables":          "Cables",
    "routers":         "Routers",
    "ethernet_switches":"Ethernet Switches",
}

def parse_extra_items(text):
    """Extract non-CCTV extra items from natural language.

    Returns (extra_items, cleaned_text) where extra_items is a list of
    (category, item_key, quantity) tuples and cleaned_text has matched
    fragments removed so the caller can parse camera info from it.
    """
    t = text.strip().lower()
    extra_items = []

    for phrase, cat, key in EXTRA_ITEM_PATTERNS:
        rgx = re.compile(r"(\d+)\s*" + re.escape(phrase) + r"s?\b")
        while True:
            m = rgx.search(t)
            if not m:
                break
            qty = int(m.group(1))
            extra_items.append((cat, key, qty))
            t = t[:m.start()] + t[m.end():]

    return extra_items, t


def parse_input(text):
    """Parse natural language like '8 cameras 2MP for mosque'"""
    t = text.strip().lower()
    count = None
    cam_type = "2mp_bullet"

    cm = re.search(r"(\d+)\s*(?:cam|unit|cctv|pcs)", t)
    if cm:
        count = int(cm.group(1))
    if not count:
        cm = re.search(r"(\d+)", t)
        if cm:
            count = int(cm.group(1))
    if count is None or count < 1:
        count = 1

    if "kpoi" in t or "kpo" in t:
        cam_type = "4mp_kpoi"
    elif "anpr" in t or "lpr" in t or "number plate" in t:
        cam_type = "4mp_anpr"
    elif "thermal" in t or "heat" in t:
        cam_type = "thermal"
    elif "ptz" in t or "pan tilt" in t or "pan-tilt" in t:
        cam_type = "6mp_ptz"
    elif "12mp" in t or "12 mp" in t:
        cam_type = "12mp_bullet"
    elif "solar" in t or "solar powered" in t or "solar 4g" in t:
        cam_type = "solar_4g_camera"
    elif "dash" in t or "dash cam" in t or "car camera" in t:
        cam_type = "dash_cam_4g"
    elif "wifi" in t or "wireless" in t or "wi-fi" in t:
        cam_type = "wifi_camera"
    elif "8mp" in t or "8 mp" in t or "4k" in t:
        cam_type = "8mp_bullet"
    elif "4mp" in t or "4 mp" in t:
        cam_type = "4mp_bullet"
    elif "2mp" in t or "2 mp" in t:
        if "eyeball" in t or "dome" in t:
            cam_type = "2mp_eyeball"
        elif "varifocal" in t or "acusense" in t or "motorized" in t:
            cam_type = "2mp_varifocal"
        else:
            cam_type = "2mp_bullet"

    # Brand-aware: if "secuview" mentioned, switch to Secuview-branded camera variant
    if "secuview" in t and not cam_type.startswith("secuview_") and not cam_type.startswith("thermal"):
        brand_key = f"secuview_{cam_type}"
        rates = load_rates()
        if brand_key in rates.get("cameras", {}):
            cam_type = brand_key

    customer = None
    for prefix in ["for ", "at ", "customer ", "client "]:
        if prefix in t:
            after = t.split(prefix, 1)[1].strip()
            words = after.split()
            name_parts = []
            for w in words:
                if w in ("moi", "with", "without", "discount", "--"):
                    break
                if re.search(r'\d', w):
                    continue
                name_parts.append(w)
            if name_parts:
                customer = " ".join(name_parts).title()
    return count, cam_type, customer

def parse_multi_input(text):
    parts = [p.strip() for p in text.split(",")]
    results = []
    for part in parts:
        if not part:
            continue
        clean = re.sub(r'\b(--moi|moi|--customer\s+\S+|--discount\s+\S+)\b', '', part, flags=re.I).strip()
        c, ct, _ = parse_input(clean)
        results.append((c, ct))
    return results

def has_multi_cameras(text):
    return "," in text

def select_nvr(cam_count, cam_type, rates, cam_list=None):
    nvrs = rates.get("nvrs", {})
    rules = rates.get("_system_rules", {})
    mapping = rules.get("nvr_selection", {}).get("mapping", [])

    # Determine NVR by channel count first
    nvr_by_count = None
    for m in mapping:
        if cam_count <= m["max_cameras"]:
            nvr_by_count = m["nvr_key"]
            break
    if nvr_by_count is None:
        nvr_by_count = "64ch_2mp"

    if cam_list:
        has_anpr = any(ct == "4mp_anpr" for _, ct in cam_list)
        has_kpoi = any(ct == "4mp_kpoi" for _, ct in cam_list)
        if has_anpr and cam_count <= 13:
            return "16ch_anpr" if "16ch_anpr" in nvrs else nvr_by_count
        if has_kpoi and cam_count <= 13:
            return "16ch_kpoi" if "16ch_kpoi" in nvrs else nvr_by_count
        return nvr_by_count

    if cam_type == "4mp_anpr" and cam_count <= 13:
        return "16ch_anpr" if "16ch_anpr" in nvrs else "16ch_2mp"
    if cam_type == "4mp_kpoi" and cam_count <= 13:
        return "16ch_kpoi" if "16ch_kpoi" in nvrs else "16ch_2mp"

    return nvr_by_count

def calc_storage(cam_count, cam_type, retention_days, rates, cam_list=None):
    """Calculate HDD requirements with RAID overhead"""
    rules = rates.get("_system_rules", {})
    gb_per_day_per_mp = rules.get("hdd_per_camera_gb", {}).get("gb_per_day_per_mp", 10)
    hdds = rates.get("hdd", {})
    nvrs = rates.get("nvrs", {})

    if cam_list:
        total_gb_per_day = 0
        for count, ct in cam_list:
            mp = 2
            for key, cam in rates.get("cameras", {}).items():
                if ct == key:
                    mp = cam.get("mp", 2)
                    break
            total_gb_per_day += count * mp * gb_per_day_per_mp
    else:
        mp = 2
        for key, cam in rates.get("cameras", {}).items():
            if cam_type == key:
                mp = cam.get("mp", 2)
                break
        total_gb_per_day = cam_count * mp * gb_per_day_per_mp
    total_gb = total_gb_per_day * retention_days
    total_tb = max(1, round(total_gb / 1000))

    parity_overhead = 0
    if retention_days >= 120:
        parity_overhead = 1  # RAID 5 = 1 disk for parity

    hdd_choice = "4tb"
    hdd_sizes = {"4tb": 4, "8tb": 8, "16tb": 16}
    if total_tb <= 4:
        hdd_choice = "4tb"
    elif total_tb <= 8:
        hdd_choice = "8tb"
    else:
        hdd_choice = "16tb"

    hdd_size = hdd_sizes[hdd_choice]
    hdd_needed = -(-total_tb // hdd_size) + parity_overhead
    return hdd_choice, max(hdd_needed, 1)

def select_switch(cam_count, rates):
    rules = rates.get("_system_rules", {}).get("switch_port_per_camera", {})
    per_16 = rules.get("cameras_per_16port", 14)
    per_8 = rules.get("cameras_per_8port", 6)
    per_24 = rules.get("cameras_per_24port", 22)
    if cam_count <= per_8:
        return "8port_poe", 1
    elif cam_count <= per_16:
        return "16port_poe", 1
    elif cam_count <= per_24:
        return "24port_poe", 1
    else:
        return "24port_poe", -(-cam_count // per_24)

def select_rack(cam_count):
    if cam_count <= 4:
        return "9u_wall"
    elif cam_count <= 16:
        return "18u"
    elif cam_count <= 32:
        return "27u"
    else:
        return "42u"

def build_quotation(count_or_list, cam_type=None, customer=None, discount_pct=0, moi_mode=False, extra_items=None):
    """Build quotation. Supports single type: build_quotation(8, '2mp_bullet', ...)
    or mixed types: build_quotation([(10,'2mp_bullet'),(3,'2mp_varifocal')], ...)"""
    rates = load_rates()

    if isinstance(count_or_list, list):
        cam_list = count_or_list
        total_count = sum(c for c, _ in cam_list)
        primary_type = max(cam_list, key=lambda x: x[0])[1] if cam_list else "2mp_bullet"
    else:
        cam_list = [(count_or_list, cam_type or "2mp_bullet")]
        total_count = count_or_list
        primary_type = cam_type or "2mp_bullet"

    retention = 120 if moi_mode else 30
    lines = []
    n = 1

    def ml(model_en, desc_en=None):
        model_ar = get_model_ar(model_en)
        desc_ar = get_model_ar(desc_en or model_en)
        return {"model": model_en, "model_ar": model_ar,
                "model_bilingual": f"{model_ar} / {model_en}",
                "desc": desc_en or model_en, "desc_ar": desc_ar,
                "desc_bilingual": f"{desc_ar} / {desc_en or model_en}"}

    cam_infos = []
    system_desc_parts = []
    for count, ct in cam_list:
        ci = rates["cameras"].get(ct, rates["cameras"]["2mp_bullet"])
        lines.append({"#": n, **ml(ci["model"], ci.get("description", ci["model"])), "qty": count, "price": ci["price"], "cat": "Cameras"})
        n += 1
        cam_infos.append({"count": count, "cam_type": ct, "cam_info": ci})
        system_desc_parts.append(f"{count} x {ci['model']}")

    # 2. NVR
    nvr_key = select_nvr(total_count, primary_type, rates, cam_list if isinstance(count_or_list, list) else None)
    nvr_info = rates["nvrs"].get(nvr_key, rates["nvrs"]["16ch_2mp"])
    lines.append({"#": n, **ml(nvr_info["model"], nvr_info.get("description", nvr_info["model"])), "qty": 1, "price": nvr_info["price"], "cat": "NVR & Recording"})
    n += 1

    # 3. HDDs with RAID if MOI
    hdd_cl = cam_list if isinstance(count_or_list, list) else None
    hdd_choice, hdd_count = calc_storage(total_count, primary_type, retention, rates, hdd_cl)
    hdd_info = rates["hdd"].get(hdd_choice, rates["hdd"]["4tb"])
    raid_note = " (RAID 5 configured)" if moi_mode else ""
    lines.append({"#": n, **ml(hdd_info["model"], f"{hdd_info['description']} x {hdd_count}{raid_note}"), "qty": hdd_count, "price": hdd_info["price"], "cat": "Storage"})
    n += 1

    # 4. PoE Switch
    sw_key, sw_count = select_switch(total_count, rates)
    sw_info = rates["switches"].get(sw_key, rates["switches"]["16port_poe"])
    lines.append({"#": n, **ml(sw_info["model"], f"{sw_info['description']} x {sw_count}"), "qty": sw_count, "price": sw_info["price"], "cat": "Networking"})
    n += 1

    # 5. Rack
    rack_key = select_rack(total_count)
    rack_info = rates["racks"].get(rack_key, rates["racks"]["18u"])
    lines.append({"#": n, **ml(rack_info["model"], rack_info["description"]), "qty": 1, "price": rack_info["price"], "cat": "Infrastructure"})
    n += 1

    # 6. Monitor
    if total_count <= 9:
        monitor_key = "24inch"
    elif total_count <= 16:
        monitor_key = "32inch"
    elif total_count <= 32:
        monitor_key = "43inch"
    else:
        monitor_key = "43inch"
    mon_info = rates["monitors"].get(monitor_key, rates["monitors"]["24inch"])
    mon_qty = 1
    lines.append({"#": n, **ml(mon_info["model"], mon_info["description"]), "qty": mon_qty, "price": mon_info["price"], "cat": "Display"})
    n += 1

    # 7. Workstation (for larger systems)
    if total_count > 8:
        ws_info = rates["workstation"]["dell_optiplex"]
        lines.append({"#": n, **ml(ws_info["model"], ws_info["description"]), "qty": 1, "price": ws_info["price"], "cat": "Workstation"})
        n += 1

    # 8. UPS
    if total_count <= 8:
        ups_key = "1kva_rack"
    elif total_count <= 24:
        ups_key = "2kva_rack"
    else:
        ups_key = "3kva_rack"
    ups_info = rates["ups"].get(ups_key, rates["ups"]["1kva_rack"])
    lines.append({"#": n, **ml(ups_info["model"], ups_info.get("description", ups_info["model"])), "qty": 1, "price": ups_info["price"], "cat": "Power"})
    n += 1
    bat_qty = 1 if total_count <= 8 else (2 if total_count <= 24 else 3)
    bat_label = ups_info.get("battery_label", "Battery Cabinet")
    lines.append({"#": n, **ml(bat_label, f"External battery for {ups_info['model']}, {bat_qty} units for minimum 1hr runtime"), "qty": bat_qty, "price": ups_info.get("battery_price", 1100), "cat": "Power"})
    n += 1

    # 9. Mounts (one line per distinct mount type)
    for ci_info in cam_infos:
        ci_count = ci_info["count"]
        ct = ci_info["cam_type"]
        ci = ci_info["cam_info"]
        mount_key = ci.get("default_mount", "wall_mount")
        if ct == "4mp_anpr":
            mnt = rates["accessories"]["pole_1_5m"]
            lines.append({"#": n, **ml(mnt["model"], f"1.5m pole mount for {ci_count} ANPR cameras"), "qty": ci_count, "price": mnt["price"], "cat": "Mounting"})
            n += 1
            mnt = rates["accessories"]["pole_mount"]
            lines.append({"#": n, **ml(mnt["model"], f"Pole mount brackets for {ci_count} cameras"), "qty": ci_count, "price": mnt["price"], "cat": "Mounting"})
            n += 1
            continue
        mnt = rates["accessories"].get(mount_key, rates["accessories"]["wall_mount"])
        label = "Hanging stands" if ct == "4mp_kpoi" else "Mounting brackets"
        lines.append({"#": n, **ml(mnt["model"], f"{label} for {ci_count} cameras"), "qty": ci_count, "price": mnt["price"], "cat": "Mounting"})
        n += 1

    # 10. Cabling
    patch_qty = total_count + 4
    lines.append({"#": n, **ml("CAT6 UTP Patch Cord", "CAT6 patch cords for connections"), "qty": patch_qty, "price": rates["accessories"]["patch_cord"]["price"], "cat": "Cabling"})
    n += 1

    # 11. Rack accessories (for larger systems)
    if total_count > 8:
        for acc_key in ["pdu", "patch_panel", "cable_manager"]:
            acc = rates["accessories"].get(acc_key, {})
            if acc:
                q = 2 if (acc_key == "patch_panel" and total_count > 24) else 1
                lines.append({"#": n, **ml(acc["model"], acc.get("description", acc["model"])), "qty": q, "price": acc["price"], "cat": "Infrastructure"})
                n += 1

    # 12. ANPR Licenses
    for _, ct in cam_list:
        if ct == "4mp_anpr":
            anpr_count = next((c for c, ct2 in cam_list if ct2 == "4mp_anpr"), 0)
            lines.append({"#": n, **ml("HIK-Central P-ANPR 1-CH License", f"ANPR license for {anpr_count} LPR cameras"), "qty": anpr_count, "price": rates["licenses"]["hikcentral_anpr_1ch"]["price"], "cat": "Licenses"})
            n += 1
            lines.append({"#": n, **ml("HIK-Central VMS Base License", "VMS base license"), "qty": 1, "price": rates["licenses"]["hikcentral_base"]["price"], "cat": "Licenses"})
            n += 1
            break

    # 13. Installation
    inst_price = rates["services"]["installation_per_camera"]["price"]
    lines.append({"#": n, **ml("Camera Installation & Cabling", f"Professional installation, cabling, conduit, configuration for {total_count} cameras"), "qty": total_count, "price": inst_price, "cat": "Services"})
    n += 1

    # 14. NVR Configuration
    lines.append({"#": n, **ml("NVR Configuration & Setup", "NVR RAID setup, network configuration, system commissioning"), "qty": 1, "price": rates["services"]["nvr_configuration"]["price"], "cat": "Services"})
    n += 1

    # 15. DSA/DIA Approvals (for MOI or large systems)
    if moi_mode or total_count > 16:
        dsa_key = "dsa_dia_approvals" if total_count > 10 else "dsa_dia_small"
        dsa_info = rates["services"].get(dsa_key, rates["services"]["dsa_dia_approvals"])
        lines.append({"#": n, **ml(dsa_info["model"], dsa_info.get("description", dsa_info["model"])), "qty": 1, "price": dsa_info["price"], "cat": "Services"})
        n += 1

    # 16. AMC (for MOI compliance)
    if moi_mode:
        amc_base = rates["services"]["amc_annual"]["price"]
        amc_per_cam = rates["services"]["amc_annual"]["price_per_camera"]
        amc_total = amc_base + (amc_per_cam * total_count)
        lines.append({"#": n, **ml("Annual Maintenance Contract (AMC)", f"12-month AMC including cleaning, firmware updates, health checks, MOI compliance, repairs ({total_count} cameras)"), "qty": 1, "price": amc_total, "cat": "Services"})
        n += 1

    # 17. Extra items (door phone, access points, speakers, etc.)
    extra_item_descs = []
    if extra_items:
        for cat, key, qty in extra_items:
            item = rates.get(cat, {}).get(key)
            if not item:
                continue
            label = EXTRA_CAT_LABEL.get(cat, cat.replace("_", " ").title())
            desc = f"{qty} x {item.get('description', item['model'])}"
            lines.append({"#": n, **ml(item["model"], desc), "qty": qty, "price": item["price"], "cat": label})
            n += 1
            extra_item_descs.append(f"{qty} x {item['model']}")

    subtotal = sum(i["qty"] * i["price"] for i in lines)
    discount_amount = round(subtotal * discount_pct / 100) if discount_pct > 0 else 0
    grand_total = subtotal - discount_amount

    all_desc_parts = system_desc_parts + extra_item_descs
    system_desc = " + ".join(all_desc_parts)

    return {
        "lines": lines, "count": total_count, "cam_type": primary_type,
        "cam_info": cam_infos[0]["cam_info"],
        "system_desc": system_desc, "cam_list": cam_list,
        "customer": customer or "Customer",
        "date": dt.date.today().isoformat(), "discount_pct": discount_pct,
        "discount_amount": discount_amount, "subtotal": subtotal,
        "grand_total": grand_total, "nvr_key": nvr_key, "hdd_choice": hdd_choice,
        "moi_mode": moi_mode, "retention_days": retention,
    }

def fn(n):
    return f"{n:,}"

def print_quotation(q):
    terms = TERMS_MOI if q["moi_mode"] else TERMS_STANDARD
    mode_tag = " [MOI COMPLIANT]" if q["moi_mode"] else ""
    print()
    print("="*80)
    print(f"  {COMPANY_NAME}{mode_tag}")
    print(f"  {COMPANY_INFO}")
    print("="*80)
    print(f"  QUOTATION")
    print(f"  Date: {q['date']}  |  Retention: {q['retention_days']} days{' (RAID 5)' if q['moi_mode'] else ''}")
    print(f"  Customer: {q['customer']}")
    sys_desc = q.get("system_desc") or f"{q['count']} x {q['cam_info']['model']}"
    print(f"  System: {sys_desc}")
    print("-"*80)
    print(f"  {'#':<3} {'Model':<38} {'Qty':<5} {'Price':<10} {'Amount':<12}")
    print("-"*80)
    for i in q["lines"]:
        amt = i["qty"] * i["price"]
        print(f"  {i['#']:<3} {i['model'][:37]:<38} {i['qty']:<5} {fn(i['price']):>8}  {fn(amt):>10}")
    print("-"*80)
    print(f"  {'Subtotal:':<55} {fn(q['subtotal']):>8} {CURRENCY}")
    if q["discount_amount"] > 0:
        print(f"  {'Discount (' + str(q['discount_pct']) + '%):':<55} {fn(q['discount_amount']):>8} {CURRENCY}")
    print(f"  {'GRAND TOTAL:':<55} {fn(q['grand_total']):>8} {CURRENCY}")
    print("="*80)
    print("  TERMS & CONDITIONS:")
    for i, t in enumerate(terms, 1):
        print(f"  {i}. {t}")
    print()
    print(f"  {COMPANY_INFO}")
    print()

def _find_item_image(item, rates):
    for ck, cv in rates.items():
        if ck.startswith("_"):
            continue
        for ik, iv in cv.items():
            if not isinstance(iv, dict):
                continue
            if iv.get("model") == item["model"] and "images" in iv:
                imgs = iv["images"]
                return imgs.get("col3") or imgs.get("col1")
    return None

def _add_logo(ws, rates):
    try:
        from openpyxl.drawing.image import Image as XlImg
        imgs_cfg = rates.get("_images", {})
        logo = imgs_cfg.get("logo")
        if logo:
            p = IMAGES_DIR / logo
            if p.exists():
                img = XlImg(str(p)); img.width = 100; img.height = 50
                img.anchor = 'A1'; ws.add_image(img); return True
            else:
                from pathlib import Path
                cwd = Path.cwd()
                alt = cwd / logo
                if alt.exists():
                    img = XlImg(str(alt)); img.width = 100; img.height = 50
                    img.anchor = 'A1'; ws.add_image(img); return True
    except Exception as e:
        import sys; print(f"[logo] Warning: could not embed logo: {e}", file=sys.stderr)
    return False

def _add_prod_img(ws, row, img_file, col=3):
    try:
        from openpyxl.drawing.image import Image as XlImg
        if img_file:
            p = IMAGES_DIR / img_file
            if p.exists():
                img = XlImg(str(p)); img.width = 55; img.height = 55
                from openpyxl.utils import get_column_letter
                img.anchor = f'{get_column_letter(col)}{row}'; ws.add_image(img); return True
            from pathlib import Path
            alt = Path.cwd() / img_file
            if alt.exists():
                img = XlImg(str(alt)); img.width = 55; img.height = 55
                from openpyxl.utils import get_column_letter
                img.anchor = f'{get_column_letter(col)}{row}'; ws.add_image(img); return True
    except Exception as e:
        import sys; print(f"[img] Warning: could not embed {img_file}: {e}", file=sys.stderr)
    return False

def export_excel(q, filename=None, arabic=False):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[excel] openpyxl required. pip install openpyxl")
        return None

    rates = load_rates()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    double_top = Border(top=Side('double'), bottom=Side('thin'))
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    tfont = Font(name='Calibri', size=18, bold=True, color="1F4E79")
    sfont = Font(name='Calibri', size=10, color="555555")
    ifont = Font(name='Calibri', size=10)
    arfont = Font(name='Calibri', size=10)
    bfont = Font(name='Calibri', size=12, bold=True)
    gfont = Font(name='Calibri', size=14, bold=True, color="1F4E79")
    tcfont = Font(name='Calibri', size=9, color="666666")
    ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
    la = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ra = Alignment(horizontal='right', vertical='center')
    ral = Alignment(horizontal='right', vertical='center', wrap_text=True)

    last_col = 9 if arabic else 7
    last_letter = get_column_letter(last_col)

    ws.merge_cells(f'A1:{last_letter}1')
    _add_logo(ws, rates)
    ws['A1'] = "  STARFOX SECURITY"
    ws['A1'].font = tfont
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 55

    ws.merge_cells(f'A2:{last_letter}2')
    subtitle = "CCTV & Smart Home Security Systems"
    sub_ar = ARABIC['subtitle']
    if q["moi_mode"]:
        subtitle += " — MOI SSD Compliant (Law No. 9/2011)"
        sub_ar += f" {ARABIC['moi_subtitle']}"
    ws_row2 = f"{subtitle} / {sub_ar}"
    ws['A2'] = ws_row2
    ws['A2'].font = sfont
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(f'A3:{last_letter}3')
    ws['A3'] = f"{COMPANY_INFO} | {ARABIC['company_info']}"
    ws['A3'].font = Font(name='Calibri', size=9, color="888888")
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(f'A5:C5')
    ws['A5'] = f"QUOTATION / {ARABIC['quotation']}"
    ws['A5'].font = Font(name='Calibri', size=14, bold=True, color="1F4E79")

    ret = f"{q['retention_days']}-day retention"
    if q["moi_mode"]:
        ret += " | RAID 5"
    date_str = f"Date / {ARABIC['date']}: {q['date']}  |  {ret}"
    ws.merge_cells(f'D5:{last_letter}5')
    ws['D5'] = date_str
    ws['D5'].font = Font(name='Calibri', size=9, color="555555")
    ws['D5'].alignment = ra

    ws.merge_cells(f'A6:{last_letter}6')
    ws['A6'] = f"Customer / {ARABIC['customer']}: {q['customer']}"
    ws['A6'].font = Font(name='Calibri', size=11, bold=True)

    sys_desc = q.get("system_desc") or f"{q['count']} x {q['cam_info']['model']}"
    ws.merge_cells(f'A7:{last_letter}7')
    ws['A7'] = f"System / {ARABIC['system']}: {sys_desc}"
    ws['A7'].font = Font(name='Calibri', size=10, color="555555")

    if arabic:
        hdrs = ["# / #", "Model / الموديل", "Description / الوصف", "Image", "Qty / الكمية", "Unit Price (QAR) / سعر الوحدة (ريال)", "Amount (QAR) / الإجمالي (ريال)"]
        widths = [6, 35, 55, 10, 10, 18, 18]
        col_offset = 0
        model_col = 2
        desc_col = 3
        model_key = "model_bilingual"
        desc_key = "desc_bilingual"
    else:
        hdrs = ["#", "Model", "Image", "Description", "Qty", "Unit Price (QAR)", "Amount (QAR)"]
        widths = [6, 28, 10, 55, 8, 16, 16]
        col_offset = 0
        model_col = 2
        desc_col = 4
        model_key = "model"
        desc_key = "desc"

    ncols = len(hdrs)
    for i, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(row=9, column=i, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = ca; c.border = thin
        ws.column_dimensions[get_column_letter(i)].width = w

    for idx, item in enumerate(q["lines"]):
        r = 10 + idx
        amt = item["qty"] * item["price"]
        ws.cell(r, 1, value=item["#"]).font = ifont; ws.cell(r, 1).alignment = ca
        ws.cell(r, model_col, value=item[model_key]).font = arfont if arabic else ifont
        if arabic:
            ws.cell(r, model_col).alignment = ral
        else:
            ws.cell(r, model_col).alignment = la
        if arabic:
            img_ref = _find_item_image(item, rates)
            _add_prod_img(ws, r, img_ref, col=4)
            ws.cell(r, desc_col, value=item[desc_key]).font = arfont; ws.cell(r, desc_col).alignment = ral
            ws.cell(r, 5, value=item["qty"]).font = ifont; ws.cell(r, 5).alignment = ca
            pc = ws.cell(r, 6, value=item["price"]); pc.font = ifont; pc.alignment = ra; pc.number_format = '#,##0'
            ac = ws.cell(r, 7, value=amt); ac.font = ifont; ac.alignment = ra; ac.number_format = '#,##0'
            for c in range(1, ncols + 1):
                ws.cell(r, c).border = thin
        else:
            img_ref = _find_item_image(item, rates)
            _add_prod_img(ws, r, img_ref, col=3)
            ws.cell(r, desc_col, value=item[desc_key]).font = ifont; ws.cell(r, desc_col).alignment = la
            ws.cell(r, 5, value=item["qty"]).font = ifont; ws.cell(r, 5).alignment = ca
            pc = ws.cell(r, 6, value=item["price"]); pc.font = ifont; pc.alignment = ra; pc.number_format = '#,##0'
            ac = ws.cell(r, 7, value=amt); ac.font = ifont; ac.alignment = ra; ac.number_format = '#,##0'
            for c in range(1, 8):
                ws.cell(r, c).border = thin
        ws.row_dimensions[r].height = 65

    tr = 10 + len(q["lines"])
    ws.merge_cells(f'A{tr}:D{tr}')
    ws.cell(tr, 1, value=f"Subtotal / {ARABIC['subtotal']}").font = bfont; ws.cell(tr, 1).alignment = ra
    ws.merge_cells(f'E{tr}:{last_letter}{tr}')
    sc = ws.cell(tr, 5, value=q["subtotal"]);
    sc.font = bfont; sc.alignment = ra; sc.number_format = '#,##0'
    for c in range(1, last_col + 1):
        ws.cell(tr, c).border = thin

    if q["discount_amount"] > 0:
        dr = tr + 1
        ws.merge_cells(f'A{dr}:D{dr}')
        ws.cell(dr, 1, value=f"Discount ({q['discount_pct']}%)").font = bfont; ws.cell(dr, 1).alignment = ra
        ws.merge_cells(f'E{dr}:{last_letter}{dr}')
        dc = ws.cell(dr, 5, value=-q["discount_amount"]); dc.font = bfont; dc.alignment = ra; dc.number_format = '#,##0'
        for c in range(1, last_col + 1):
            ws.cell(dr, c).border = thin
        gr = dr + 1
    else:
        gr = tr + 1

    ws.merge_cells(f'A{gr}:D{gr}')
    ws.cell(gr, 1, value=f"GRAND TOTAL / {ARABIC['grand_total']}").font = gfont; ws.cell(gr, 1).alignment = ra
    ws.merge_cells(f'E{gr}:{last_letter}{gr}')
    gc = ws.cell(gr, 5, value=q["grand_total"]); gc.font = gfont; gc.alignment = ra; gc.number_format = '#,##0'
    for c in range(1, last_col + 1):
        ws.cell(gr, c).border = double_top

    r2 = gr + 2
    ws.merge_cells(f'A{r2}:{last_letter}{r2}')
    ws.cell(r2, 1, value=f"TERMS & CONDITIONS / {ARABIC['terms_title']}").font = Font(name='Calibri', size=10, bold=True)
    terms = TERMS_MOI if q["moi_mode"] else TERMS_STANDARD
    terms_ar = ARABIC['terms_moi'] if q["moi_mode"] else ARABIC['terms_standard']
    for i, (t, ta) in enumerate(zip(terms, terms_ar)):
        r = r2 + 1 + i
        ws.merge_cells(f'A{r}:{last_letter}{r}')
        ws.cell(r, 1, value=f"{i+1}. {t} / {ta}").font = tcfont; ws.cell(r, 1).alignment = la

    sr = r2 + len(terms) + 2
    ws.merge_cells(f'A{sr}:{last_letter}{sr}')
    ws[f'A{sr}'] = f"{COMPANY_INFO} | {ARABIC['company_info']}"
    ws[f'A{sr}'].font = Font(name='Calibri', size=9, color="888888")
    ws[f'A{sr}'].alignment = Alignment(horizontal='center')

    if filename is None:
        ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
        safe = re.sub(r'[^\w\s-]', '', q['customer'])[:20].strip()
        mode = "MOI-" if q["moi_mode"] else ""
        ar = "AR-" if arabic else ""
        filename = QUOTES_DIR / f"quotation-{ar}{mode}{safe}-{ts}.xlsx"

    wb.save(str(filename))
    print(f"[excel] Quotation saved: {filename}")
    return str(filename)

def update_rates_from_excel(filepath):
    try:
        import openpyxl
    except ImportError:
        print("[error] pip install openpyxl"); return False
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rates = load_rates()
    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 3:
            continue
        cat = str(row[0]).strip().lower() if row[0] else ""
        key = str(row[1]).strip().lower() if row[1] else ""
        new_price = row[2]
        if not cat or not key or new_price is None:
            continue
        try:
            new_price = float(new_price)
        except (ValueError, TypeError):
            continue
        rcat = next((rc for rc in rates if rc.startswith("_") is False and (cat == rc or cat in rc)), None)
        if not rcat:
            continue
        ritem = next((rk for rk, rv in rates[rcat].items() if key == rk or key in rk or key in rv.get("model", "").lower()), None)
        if not ritem:
            continue
        old = rates[rcat][ritem]["price"]
        rates[rcat][ritem]["price"] = new_price
        updated += 1
        print(f"  {rcat}.{ritem}: {old} -> {new_price} {CURRENCY}")
    save_rates(rates)
    print(f"[import] Updated {updated} items")
    return True

def interactive_mode():
    print(); print("="*50); print("  STARFOX AI Quotation Maker v2"); print("="*50)
    print("  Commands:")
    print("  - '8 cameras 2MP'  (standard)")
    print("  - '8 cameras 2MP --moi'  (MOI compliant)")
    print("  - '10 cameras 2mp, 3 varifocal, 1 kpoi'  (mixed types)")
    print("  - 'rates' or 'list'  (show rate card)")
    print("  - '--arabic' for bilingual EN/AR output")
    print("  - 'exit' to quit"); print()
    moi = False; arabic = False
    while True:
        try:
            inp = input("you: ").strip()
        except (EOFError, KeyboardInterrupt):
            print(); break
        if not inp: continue
        if inp.lower() in ("exit", "quit", "bye"): print("Goodbye!"); break
        if inp.lower() in ("list", "rates", "list rates"): list_rates(); continue
        if "--arabic" in inp.lower():
            arabic = True
            inp = inp.replace("--arabic", "").strip()
        if "--moi" in inp.lower() or "/moi" in inp:
            moi = True
            inp = inp.replace("--moi", "").replace("/moi", "").strip()
        extra_items, cleaned = parse_extra_items(inp)
        if has_multi_cameras(cleaned):
            cam_list = parse_multi_input(cleaned)
            q = build_quotation(cam_list, customer=None, moi_mode=moi, extra_items=extra_items)
        else:
            count, cam_type, customer = parse_input(cleaned)
            q = build_quotation(count, cam_type, customer, moi_mode=moi, extra_items=extra_items)
        q["arabic"] = arabic
        print_quotation(q)
        export_excel(q, arabic=arabic)
        moi = False; arabic = False

def main():
    if "--list-rates" in sys.argv:
        list_rates(); return
    if "--update-rates" in sys.argv:
        idx = sys.argv.index("--update-rates")
        if idx + 1 < len(sys.argv):
            update_rates_from_excel(sys.argv[idx + 1])
        else:
            print("Usage: --update-rates <path_to_excel.xlsx>")
        return
    if "--interactive" in sys.argv:
        interactive_mode(); return
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    moi_mode = "--moi" in sys.argv
    arabic_mode = "--arabic" in sys.argv
    if not args:
        interactive_mode(); return
    text = args[0]
    extra_items, cleaned_text = parse_extra_items(text)
    customer = None; discount_pct = 0
    if "--customer" in sys.argv:
        ci = sys.argv.index("--customer")
        if ci + 1 < len(sys.argv): customer = sys.argv[ci + 1]
    if "--discount" in sys.argv:
        di = sys.argv.index("--discount")
        if di + 1 < len(sys.argv):
            try: discount_pct = float(sys.argv[di + 1])
            except ValueError: pass
    if has_multi_cameras(cleaned_text):
        cam_list = parse_multi_input(cleaned_text)
        q = build_quotation(cam_list, customer=customer, discount_pct=discount_pct, moi_mode=moi_mode, extra_items=extra_items)
    else:
        count, cam_type, parsed_customer = parse_input(cleaned_text)
        if customer is None: customer = parsed_customer
        q = build_quotation(count, cam_type, customer, discount_pct, moi_mode, extra_items=extra_items)
    q["arabic"] = arabic_mode
    print_quotation(q)
    export_excel(q, arabic=arabic_mode)
    mode_str = f"{'MOI ' if moi_mode else ''}{'AR ' if arabic_mode else ''}"
    extra_count = sum(qty for _, _, qty in extra_items)
    print(f"\n[info] {q['count']} cameras | {extra_count} extra items | {mode_str}mode")
    print(f"[info] System: {q.get('system_desc', q['cam_type'])}")
    print(f"[info] Grand Total: {fn(q['grand_total'])} {CURRENCY}")

if __name__ == "__main__":
    main()
