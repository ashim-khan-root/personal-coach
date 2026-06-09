"""
STARFOX AI Quotation Maker - CCTV & Smart Home Security Systems

Usage:
  python coach/tools/make_quotation.py "8 cameras 2MP"
  python coach/tools/make_quotation.py "3 cameras 4MP KPOI for villa"
  python coach/tools/make_quotation.py "12 cameras 2MP eyeball" --customer "Al Thani Villa" --discount 5
  python coach/tools/make_quotation.py --update-rates path/to/new_rates.xlsx
  python coach/tools/make_quotation.py --list-rates
  python coach/tools/make_quotation.py --interactive
"""
import sys, json, re, datetime as dt
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
MEM_DIR = BASE / "memory"
RATES_PATH = MEM_DIR / "rates.json"
QUOTES_DIR = MEM_DIR / "quotations"
QUOTES_DIR.mkdir(parents=True, exist_ok=True)

CURRENCY = "QAR"
COMPANY_NAME = "STARFOX SECURITY"
COMPANY_INFO = "Salwa Road, Doha, Qatar | Tel: +974 44691741 | info@starfoxsecu.com | www.starfoxsecu.com"
IMAGES_DIR = BASE / "memory" / "images"
TERMS = [
    "All the above prices are in QAR (Qatari Riyals)",
    "Terms of payment: 50% upon signing the contract as downpayment and 50% before installation and handover",
    "Warranty: 2 YEARS for NVR and Cameras",
    "Any changes or addition of cameras as recommended by MOI will be quoted separately",
    "Any civil, mechanical, and electrical work if any will be charged accordingly",
    "Materials prices may vary at the time of installation due to current regional situation",
    "This quotation is subject to STARFOX standard terms and conditions which are available on request",
    "Price and specifications may change without prior notice",
]

def load_rates():
    with open(RATES_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def save_rates(rates):
    with open(RATES_PATH, "w", encoding="utf-8") as f:
        json.dump(rates, f, ensure_ascii=False, indent=2)
    print(f"[rates] Saved to {RATES_PATH}")

def list_rates():
    rates = load_rates()
    print("=" * 60)
    print(f"  STARFOX RATE CARD ({CURRENCY})")
    print("=" * 60)
    for category, items in rates.items():
        if category.startswith("_"):
            continue
        print(f"\n  [{category.upper()}]")
        for key, item in items.items():
            price = item.get("price", 0)
            model = item.get("model", key)
            print(f"    {key:30s} {price:>8,} {CURRENCY}  | {model}")
    print()

def parse_input(text):
    """Parse natural language like '8 cameras 2MP' or '8 cameras 2MP and 2 cameras 4MP varifocal'"""
    text = text.strip().lower()
    customer = None
    c_match = re.search(r"(?:for|at|customer|client)\s+['\"]?([A-Za-z0-9\s]+?)['\"]?(?:\s+\d|$)", text)
    if c_match:
        customer = c_match.group(1).strip()

    parts = re.split(r'\s+(?:and|&|\+)\s+|\s*,\s*', text)
    cameras = []
    for part in parts:
        part = part.strip()
        if not part or any(w in part for w in ['for', 'at', 'customer', 'client']):
            continue

        count = None
        cam_type = "2mp_bullet"
        count_match = re.search(r"(\d+)\s*(?:x|cam|unit|cctv|no|pcs|pc)", part)
        if count_match:
            count = int(count_match.group(1))
        if not count:
            count_match = re.search(r"(\d+)", part)
            if count_match:
                count = int(count_match.group(1))
        if count is None or count < 1:
            count = 1

        if "8mp" in part or "8 mp" in part:
            cam_type = "2mp_bullet"
        elif "4mp" in part or "4 mp" in part:
            if "varifocal" in part or "motorized" in part or "acusense" in part:
                cam_type = "4mp_varifocal"
            elif "kpoi" in part or "kpo" in part:
                cam_type = "4mp_kpoi"
            elif "anpr" in part or "lpr" in part or "number plate" in part:
                cam_type = "4mp_anpr"
            else:
                cam_type = "4mp_kpoi"
        elif "2mp" in part or "2 mp" in part:
            if "eyeball" in part or "dome" in part:
                cam_type = "2mp_eyeball"
            elif "varifocal" in part or "acusense" in part or "motorized" in part:
                cam_type = "2mp_varifocal"
            else:
                cam_type = "2mp_bullet"
        cameras.append((count, cam_type))
    if not cameras:
        cameras = [(1, "2mp_bullet")]
    return cameras, customer

def select_nvr(cam_count, cameras, rates):
    """Auto-select appropriate NVR based on total camera count and types"""
    nvrs = rates.get("nvrs", {})
    rules = rates.get("_system_rules", {})
    nvr_map = rules.get("nvr_selection", {}).get("mapping", [])

    highest_type = max(cameras, key=lambda x: _cam_res_order(x[1]))[1]
    if highest_type == "4mp_anpr":
        if cam_count <= 16 and "16ch_anpr" in nvrs:
            return "16ch_anpr"
    elif highest_type == "4mp_kpoi":
        if cam_count <= 16 and "16ch_kpoi" in nvrs:
            return "16ch_kpoi"
    for mapping in nvr_map:
        if cam_count <= mapping["max_cameras"]:
            return mapping["nvr_key"]
    return "64ch_2mp"

def _cam_res_order(cam_type):
    order = {"2mp_bullet": 1, "2mp_eyeball": 1, "2mp_varifocal": 1,
             "4mp_kpoi": 2, "4mp_varifocal": 2, "4mp_anpr": 2}
    return order.get(cam_type, 0)

def select_hdd(cameras, nvr_key, rates):
    """Calculate HDD requirements for 30-day retention across mixed camera types"""
    nvrs = rates.get("nvrs", {})
    nvr_info = nvrs.get(nvr_key, {})
    max_hdds = nvr_info.get("max_hdds", 2)
    total_storage_gb = 0
    for count, cam_type in cameras:
        mp = 2
        if "4mp" in cam_type:
            mp = 4
        elif "8mp" in cam_type:
            mp = 8
        total_storage_gb += count * mp * 10 * 30
    total_tb_needed = max(1, round(total_storage_gb / 1000))
    hdd_choice = "4tb"
    if total_tb_needed <= 4:
        hdd_choice = "4tb"
    elif total_tb_needed <= 8:
        hdd_choice = "8tb"
    else:
        hdd_choice = "16tb"
    hdd_size = {"4tb": 4, "8tb": 8, "16tb": 16}[hdd_choice]
    hdd_count = min(max(1, -(-total_tb_needed // hdd_size)), max_hdds)
    return hdd_choice, hdd_count

def select_switch(cam_count, rates):
    """Auto-select PoE switch based on camera count"""
    rules = rates.get("_system_rules", {}).get("switch_port_per_camera", {})
    cams_per_16 = rules.get("cameras_per_16port", 14)
    cams_per_8 = rules.get("cameras_per_8port", 6)
    if cam_count <= cams_per_8:
        return "8port_poe", 1
    elif cam_count <= cams_per_16:
        return "16port_poe", 1
    else:
        sw_count = -(-cam_count // cams_per_16)
        return "16port_poe", sw_count

def select_rack(cam_count, rates):
    """Auto-select rack based on system size"""
    return ("27u", True) if cam_count > 8 else ("18u", False)

def build_quotation(cameras, customer=None, discount_pct=0):
    """Generate full quotation line items from list of (count, cam_type) pairs"""
    rates = load_rates()
    cameras_cfg = rates.get("cameras", {})
    accessories = rates.get("accessories", {})
    services = rates.get("services", {})
    ups_data = rates.get("ups", {})
    switches_cfg = rates.get("switches", {})
    total_count = sum(c for c, _ in cameras)
    lines = []
    line_no = 1

    for count, cam_type in cameras:
        cam_info = cameras_cfg.get(cam_type, cameras_cfg.get("2mp_bullet"))
        lines.append({"#": line_no, "model": cam_info["model"], "desc": cam_info["description"], "qty": count, "price": cam_info["price"], "cat": "Camera"})
        line_no += 1

    nvr_key = select_nvr(total_count, cameras, rates)
    nvr_info = rates["nvrs"].get(nvr_key, {})
    lines.append({"#": line_no, "model": nvr_info.get("model", "NVR"), "desc": nvr_info.get("description", ""), "qty": 1, "price": nvr_info["price"], "cat": "NVR"})
    line_no += 1

    hdd_choice, hdd_count = select_hdd(cameras, nvr_key, rates)
    hdd_info = rates["hdd"].get(hdd_choice, rates["hdd"]["4tb"])
    lines.append({"#": line_no, "model": hdd_info["model"], "desc": f"{hdd_info['description']} x {hdd_count}", "qty": hdd_count, "price": hdd_info["price"], "cat": "Storage"})
    line_no += 1

    sw_key, sw_count = select_switch(total_count, rates)
    sw_info = switches_cfg.get(sw_key, switches_cfg.get("16port_poe"))
    lines.append({"#": line_no, "model": sw_info["model"], "desc": f"{sw_info['description']} x {sw_count}", "qty": sw_count, "price": sw_info["price"], "cat": "Networking"})
    line_no += 1

    rack_key, _ = select_rack(total_count, rates)
    rack_info = rates["racks"].get(rack_key, rates["racks"]["18u"])
    lines.append({"#": line_no, "model": rack_info["model"], "desc": rack_info["description"], "qty": 1, "price": rack_info["price"], "cat": "Infrastructure"})
    line_no += 1

    lines.append({"#": line_no, "model": "23.8 inch LED Monitor", "desc": "Full HD monitor for surveillance viewing", "qty": 1, "price": 445, "cat": "Display"})
    line_no += 1

    ups_key = "2kva_rack" if total_count > 8 else "1kva_rack"
    ups_entry = ups_data.get(ups_key, {})
    bat_count = 2 if total_count > 8 else 1
    lines.append({"#": line_no, "model": ups_entry.get("model", "UPS"), "desc": ups_entry.get("description", ""), "qty": 1, "price": ups_entry.get("price", 1200), "cat": "Power"})
    line_no += 1

    bat_label = "48V Battery Cabinet" if total_count > 8 else f"Battery for {ups_entry.get('model', '1KVA UPS')}"
    bat_desc = f"48V battery cabinet with built-in {bat_count * 4} batteries for 2KVA UPS" if total_count > 8 else f"Battery for {ups_entry.get('model', '1KVA UPS')}"
    lines.append({"#": line_no, "model": bat_label, "desc": bat_desc, "qty": bat_count, "price": ups_entry.get("battery_price", 900), "cat": "Power"})
    line_no += 1

    mount_count = 0
    for count, cam_type in cameras:
        cam_info = cameras_cfg.get(cam_type, cameras_cfg.get("2mp_bullet"))
        mount_key = cam_info.get("default_mount", "wall_mount")
        mount_info = accessories.get(mount_key, accessories.get("wall_mount"))
        lines.append({"#": line_no, "model": mount_info["model"], "desc": f"Mount for {cam_info['model']} x {count}", "qty": count, "price": mount_info["price"], "cat": "Accessories"})
        line_no += 1
        mount_count += count

    lines.append({"#": line_no, "model": "Patch Cord CAT6", "desc": "CAT6 patch cords for connections", "qty": total_count + 4, "price": 7, "cat": "Accessories"})
    line_no += 1

    if total_count > 8:
        lines.append({"#": line_no, "model": "PDU for Rack", "desc": "Power Distribution Unit for rack", "qty": 1, "price": 65, "cat": "Infrastructure"})
        line_no += 1
        lines.append({"#": line_no, "model": "Patch Panel 24-Port", "desc": "24-port CAT6 patch panel", "qty": 1, "price": 145, "cat": "Infrastructure"})
        line_no += 1
        lines.append({"#": line_no, "model": "Cable Manager", "desc": "1U horizontal cable manager", "qty": 1, "price": 50, "cat": "Infrastructure"})
        line_no += 1

    inst_price = services.get("installation_per_camera", {}).get("price", 400)
    lines.append({"#": line_no, "model": "Camera System Installation", "desc": f"Installation, configuration, piping, cabling for {total_count} cameras", "qty": total_count, "price": inst_price, "cat": "Services"})
    line_no += 1

    subtotal = sum(item["qty"] * item["price"] for item in lines)
    discount_amount = round(subtotal * discount_pct / 100) if discount_pct > 0 else 0
    grand_total = subtotal - discount_amount

    cam_desc = " + ".join(f"{c} x {cameras_cfg.get(t, {}).get('model',t)}" for c, t in cameras)
    primary_type = cameras[0][1] if cameras else "2mp_bullet"

    return {
        "lines": lines, "cameras": cameras, "total_count": total_count,
        "cam_desc": cam_desc, "customer": customer or "Customer",
        "date": dt.date.today().isoformat(), "discount_pct": discount_pct,
        "discount_amount": discount_amount, "subtotal": subtotal,
        "grand_total": grand_total, "nvr_key": nvr_key, "hdd_choice": hdd_choice,
    }

def fn(n):
    return f"{n:,}"

def print_quotation(q):
    """Print formatted quotation to console"""
    print()
    print("=" * 80)
    print(f"  {COMPANY_NAME}")
    print(f"  {COMPANY_INFO}")
    print("=" * 80)
    print(f"  QUOTATION")
    print(f"  Date: {q['date']}")
    print(f"  Customer: {q['customer']}")
    print(f"  System: {q['cam_desc']}")
    print("-" * 80)
    print(f"  {'#':<3} {'Model':<35} {'Qty':<5} {'Price':<10} {'Amount':<12}")
    print("-" * 80)
    for item in q["lines"]:
        amt = item["qty"] * item["price"]
        print(f"  {item['#']:<3} {item['model'][:34]:<35} {item['qty']:<5} {fn(item['price']):>8}  {fn(amt):>10}")
    print("-" * 80)
    print(f"  {'Subtotal:':<50} {fn(q['subtotal']):>8} {CURRENCY}")
    if q['discount_amount'] > 0:
        print(f"  {'Discount (' + str(q['discount_pct']) + '%):':<50} {fn(q['discount_amount']):>8} {CURRENCY}")
    print(f"  {'GRAND TOTAL:':<50} {fn(q['grand_total']):>8} {CURRENCY}")
    print("=" * 80)
    print("  TERMS & CONDITIONS:")
    for i, term in enumerate(TERMS, 1):
        print(f"  {i}. {term}")
    print()
    print(f"  {COMPANY_INFO}")
    print()

def _find_item_image(item, rates):
    """Find product image from rates.json by matching model name"""
    for ck, cv in rates.items():
        if ck.startswith("_"):
            continue
        for ik, iv in cv.items():
            if not isinstance(iv, dict):
                continue
            if iv.get("model") == item["model"] and "images" in iv:
                imgs = iv["images"]
                return imgs.get("col3") or imgs.get("col1")
    return None

def _add_logo(ws, rates):
    try:
        from openpyxl.drawing.image import Image as XlImg
        imgs_cfg = rates.get("_images", {})
        logo = imgs_cfg.get("logo")
        if logo:
            p = IMAGES_DIR / logo
            if p.exists():
                img = XlImg(str(p)); img.width = 100; img.height = 50
                img.anchor = 'A1'; ws.add_image(img); return True
            else:
                from pathlib import Path
                cwd = Path.cwd()
                alt = cwd / logo
                if alt.exists():
                    img = XlImg(str(alt)); img.width = 100; img.height = 50
                    img.anchor = 'A1'; ws.add_image(img); return True
    except Exception as e:
        import sys; print(f"[logo] Warning: could not embed logo: {e}", file=sys.stderr)
    return False

def _add_prod_img(ws, row, img_file, col=3):
    try:
        from openpyxl.drawing.image import Image as XlImg
        if img_file:
            p = IMAGES_DIR / img_file
            if p.exists():
                img = XlImg(str(p)); img.width = 55; img.height = 55
                from openpyxl.utils import get_column_letter
                img.anchor = f'{get_column_letter(col)}{row}'; ws.add_image(img); return True
            else:
                alt = __import__('pathlib').Path.cwd() / img_file
                if alt.exists():
                    img = XlImg(str(alt)); img.width = 55; img.height = 55
                    from openpyxl.utils import get_column_letter
                    img.anchor = f'{get_column_letter(col)}{row}'; ws.add_image(img); return True
    except Exception as e:
        import sys; print(f"[img] Warning: could not embed {img_file}: {e}", file=sys.stderr)
    return False

def export_excel(q, filename=None):
    """Generate formatted quotation Excel with product images"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[excel] openpyxl not installed. Run: pip install openpyxl")
        return None

    rates = load_rates()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    tfont = Font(name='Calibri', size=18, bold=True, color="1F4E79")
    sfont = Font(name='Calibri', size=10, color="555555")
    ifont = Font(name='Calibri', size=10)
    bfont = Font(name='Calibri', size=12, bold=True)
    gfont = Font(name='Calibri', size=14, bold=True, color="1F4E79")
    tcfont = Font(name='Calibri', size=9, color="666666")
    ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
    la = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ra = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('A1:G1')
    _add_logo(ws, rates)
    ws.row_dimensions[1].height = 55

    ws.merge_cells('A2:G2')
    ws['A2'] = "STARFOX SECURITY"
    ws['A2'].font = tfont
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A3:G3')
    ws['A3'] = "CCTV & Smart Home Security Systems"
    ws['A3'].font = sfont
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A4:G4')
    ws['A4'] = COMPANY_INFO
    ws['A4'].font = Font(name='Calibri', size=9, color="888888")
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A6:C6')
    ws['A6'] = "QUOTATION"
    ws['A6'].font = Font(name='Calibri', size=14, bold=True, color="1F4E79")

    ws.merge_cells('D6:G6')
    ws['D6'] = f"Date: {q['date']}"
    ws['D6'].font = Font(name='Calibri', size=10)
    ws['D6'].alignment = ra

    ws.merge_cells('A7:G7')
    ws['A7'] = f"Customer: {q['customer']}"
    ws['A7'].font = Font(name='Calibri', size=10, bold=True)

    ws.merge_cells('A8:G8')
    ws['A8'] = f"System: {q['cam_desc']}"
    ws['A8'].font = Font(name='Calibri', size=10, color="555555")

    headers = ["#", "Model", "Image", "Description", "Qty", "Unit Price (QAR)", "Amount (QAR)"]
    col_widths = [6, 28, 10, 50, 8, 16, 16]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=9, column=i, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = ca; cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = w

    for idx, item in enumerate(q["lines"]):
        row = 10 + idx
        amt = item["qty"] * item["price"]
        ws.cell(row=row, column=1, value=item["#"]).font = ifont
        ws.cell(row=row, column=1).alignment = ca
        ws.cell(row=row, column=2, value=item["model"]).font = ifont
        ws.cell(row=row, column=2).alignment = la
        img_ref = _find_item_image(item, rates)
        _add_prod_img(ws, row, img_ref, col=3)
        ws.cell(row=row, column=4, value=item["desc"]).font = ifont
        ws.cell(row=row, column=4).alignment = la
        ws.cell(row=row, column=5, value=item["qty"]).font = ifont
        ws.cell(row=row, column=5).alignment = ca
        pc = ws.cell(row=row, column=6, value=item["price"])
        pc.font = ifont; pc.alignment = ra; pc.number_format = '#,##0'
        ac = ws.cell(row=row, column=7, value=amt)
        ac.font = ifont; ac.alignment = ra; ac.number_format = '#,##0'
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = thin_border
        ws.row_dimensions[row].height = 65

    tr = 10 + len(q["lines"])
    ws.merge_cells(f'A{tr}:E{tr}')
    ws.cell(row=tr, column=1, value="Subtotal").font = bfont
    ws.cell(row=tr, column=1).alignment = ra
    ws.merge_cells(f'F{tr}:G{tr}')
    sc = ws.cell(row=tr, column=6, value=q["subtotal"])
    sc.font = bfont; sc.alignment = ra; sc.number_format = '#,##0'
    for c in range(1, 8):
        ws.cell(row=tr, column=c).border = thin_border

    if q["discount_amount"] > 0:
        dr = tr + 1
        ws.merge_cells(f'A{dr}:E{dr}')
        ws.cell(row=dr, column=1, value=f"Discount ({q['discount_pct']}%)").font = bfont
        ws.cell(row=dr, column=1).alignment = ra
        ws.merge_cells(f'F{dr}:G{dr}')
        dc = ws.cell(row=dr, column=6, value=-q["discount_amount"])
        dc.font = bfont; dc.alignment = ra; dc.number_format = '#,##0'
        for c in range(1, 8):
            ws.cell(row=dr, column=c).border = thin_border
        gr = dr + 1
    else:
        gr = tr + 1

    ws.merge_cells(f'A{gr}:E{gr}')
    ws.cell(row=gr, column=1, value="GRAND TOTAL").font = gfont
    ws.cell(row=gr, column=1).alignment = ra
    ws.merge_cells(f'F{gr}:G{gr}')
    gc = ws.cell(row=gr, column=6, value=q["grand_total"])
    gc.font = gfont; gc.alignment = ra; gc.number_format = '#,##0'
    for c in range(1, 8):
        ws.cell(row=gr, column=c).border = Border(top=Side('double'), bottom=Side('double'))

    tr2 = gr + 2
    ws.merge_cells(f'A{tr2}:G{tr2}')
    ws.cell(row=tr2, column=1, value="TERMS & CONDITIONS:").font = Font(name='Calibri', size=10, bold=True)
    for i, term in enumerate(TERMS):
        r = tr2 + 1 + i
        ws.merge_cells(f'A{r}:G{r}')
        ws.cell(row=r, column=1, value=f"{i+1}. {term}").font = tcfont
        ws.cell(row=r, column=1).alignment = la

    sig_row = tr2 + len(TERMS) + 2
    ws.merge_cells(f'A{sig_row}:G{sig_row}')
    ws[f'A{sig_row}'] = COMPANY_INFO
    ws[f'A{sig_row}'].font = Font(name='Calibri', size=9, color="888888")
    ws[f'A{sig_row}'].alignment = Alignment(horizontal='center')

    if filename is None:
        ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
        safe = re.sub(r'[^\w\s-]', '', q['customer']).strip()[:20]
        filename = QUOTES_DIR / f"quotation-{safe}-{ts}.xlsx"

    wb.save(str(filename))
    print(f"[excel] Quotation saved: {filename}")
    return str(filename)

def update_rates_from_excel(filepath):
    """Import new prices from an Excel rate sheet into rates.json"""
    try:
        import openpyxl
    except ImportError:
        print("[error] openpyxl required. pip install openpyxl")
        return False
    print(f"[import] Reading rates from: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rates = load_rates()
    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 3:
            continue
        cat_key = str(row[0]).strip().lower() if row[0] else ""
        item_key = str(row[1]).strip().lower() if row[1] else ""
        new_price = row[2]
        if not cat_key or not item_key or new_price is None:
            continue
        try:
            new_price = float(new_price)
        except (ValueError, TypeError):
            continue
        cat = None
        for rcat, ritems in rates.items():
            if rcat.startswith("_"):
                continue
            if cat_key == rcat or cat_key in rcat:
                cat = rcat
                break
        if not cat:
            continue
        item = None
        for rkey, ritem in rates[cat].items():
            if item_key == rkey:
                item = rkey
                break
            if item_key in rkey or item_key in ritem.get("model", "").lower():
                item = rkey
                break
        if not item:
            continue
        old_price = rates[cat][item]["price"]
        rates[cat][item]["price"] = new_price
        updated += 1
        print(f"  Updated {cat}.{item}: {old_price} -> {new_price} {CURRENCY}")
    save_rates(rates)
    print(f"[import] Updated {updated} items")
    return True

def interactive_mode():
    """Run in interactive chat-like mode"""
    print()
    print("=" * 50)
    print("  STARFOX AI Quotation Maker")
    print("=" * 50)
    print("  Tell me what you need, e.g.:")
    print("  - '8 cameras 2MP'")
    print("  - '3 cameras 4MP KPOI'")
    print("  - 'list rates'")
    print("  - 'exit' to quit")
    print()
    while True:
        try:
            inp = input("you: ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            break
        if not inp:
            continue
        if inp.lower() in ("exit", "quit", "bye"):
            print("Goodbye!")
            break
        if inp.lower() in ("list", "rates", "list rates"):
            list_rates()
            continue
        cameras, customer = parse_input(inp)
        q = build_quotation(cameras, customer)
        print_quotation(q)
        export_excel(q)

def main():
    if "--list-rates" in sys.argv:
        list_rates()
        return
    if "--update-rates" in sys.argv:
        idx = sys.argv.index("--update-rates")
        if idx + 1 < len(sys.argv):
            update_rates_from_excel(sys.argv[idx + 1])
        else:
            print("Usage: --update-rates <path_to_excel.xlsx>")
        return
    if "--interactive" in sys.argv:
        interactive_mode()
        return
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        interactive_mode()
        return
    text = args[0]
    customer = None
    discount_pct = 0
    if "--customer" in sys.argv:
        ci = sys.argv.index("--customer")
        if ci + 1 < len(sys.argv):
            customer = sys.argv[ci + 1]
    if "--discount" in sys.argv:
        di = sys.argv.index("--discount")
        if di + 1 < len(sys.argv):
            try:
                discount_pct = float(sys.argv[di + 1])
            except ValueError:
                pass
    cameras, parsed_customer = parse_input(text)
    if customer is None:
        customer = parsed_customer
    q = build_quotation(cameras, customer, discount_pct)
    print_quotation(q)
    export_excel(q)
    cam_types = ", ".join(f"{c}x {t}" for c, t in cameras)
    print(f"\n[info] Quotation generated: {cam_types}")
    if q["discount_amount"] > 0:
        print(f"[info] Discount: {q['discount_pct']}% = {fn(q['discount_amount'])} {CURRENCY}")
    print(f"[info] Grand Total: {fn(q['grand_total'])} {CURRENCY}")

if __name__ == "__main__":
    main()
